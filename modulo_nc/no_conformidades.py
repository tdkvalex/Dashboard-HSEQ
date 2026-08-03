#!/usr/bin/env python3
"""
Procesa el registro de No Conformidades y escribe datos_nc.json, que alimenta el
dashboard del módulo y la portada de la Suite QAQC.

Uso (cada corte, con el archivo del período):

    python3 no_conformidades.py --data /ruta/Data_NCR.xlsx

Opcional:
    --hoy 2026-07-27     fecha de referencia para antigüedad y atrasos
                         (por defecto, el día en que se corre el script)

------------------------------------------------------------------------------
DE DÓNDE SALE CADA DATO
------------------------------------------------------------------------------
Hoja «Observaciones», encabezado fila 1, datos fila 2+:

    col A  Nombre Del Proyecto      col D  Tipo De Emisión (interna/externa)
    col F  Tipo (NC / PNC / …)      col I  Especialidad
    col J  Persona Asignada         col L  Fecha De Creación
    col M  Título                   col T  Costo en UF
    col V  Fecha De Cierre          col Y  Estatus

------------------------------------------------------------------------------
REGLAS DE CLASIFICACIÓN
------------------------------------------------------------------------------
ORIGEN     «Interna BSMT» -> Interna · «Externa Cliente» y «Externa Subcontrato»
           -> Externa (se guarda además el detalle de cuál de las dos).

ESTADO     «Cerrado» -> Cerrada · el resto (Iniciado, Listo para revisión,
           No aceptado) -> Abierta. Se conserva el estatus original, porque
           «No aceptado» no es lo mismo que «Iniciado» para gestionar.

ATRASO     Hay 10 días para responder una NC desde que se emite; del día 11 en
           adelante corre atraso (PLAZO_RESPUESTA). Se cuenta sobre la FECHA DE
           EMISIÓN, que las dos fuentes traen, y no sobre una fecha comprometida
           de cierre, que NINGUNA trae —en el registro principal la columna
           «Fecha De Cierre» solo se llena al cerrar, y la planilla del cliente
           tampoco declara plazo—. Por eso el atraso sí es calculable.
               atrasada       = abierta con más de 10 días desde la emisión
               días de atraso = antigüedad − 10
           Se sigue reportando la ANTIGÜEDAD de cada abierta, que es lo que
           permite priorizar entre las atrasadas.

TIEMPO DE  Solo para las cerradas: días entre la fecha de creación y la de
CIERRE     cierre. Es el indicador de qué tan rápido reacciona cada proyecto.

EXCLUIDOS  Las **opciones de mejora** no entran al módulo: son propuestas, no
           hallazgos que haya que corregir, y mezclarlas infla el universo y
           ensucia el % de cierre con algo que nadie está obligado a cerrar.
           Se cuentan y se declaran en «Control de calidad del dato», para que
           la exclusión se vea. La constante es TIPOS_EXCLUIDOS.
"""

import argparse
import json
import sys
import unicodedata
import warnings
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from math import floor
from pathlib import Path
from statistics import median

warnings.filterwarnings("ignore")

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Falta openpyxl.  Instálalo con:  pip install openpyxl")

AQUI = Path(__file__).resolve().parent

MODULO = {
    "id": "NC",
    "nombre": "No Conformidades",
    "descripcion": "Hallazgos de calidad y su corrección",
    "cliente": "Besalco Montajes",
}

# El registro nombra los proyectos con su código; la suite los conoce por su
# nombre corto. Oficina Central no es un proyecto de obra pero sí genera NC,
# así que se reporta aparte y no se mezcla con los tres proyectos.
# Código, nombre y cliente son exactamente los que usa el módulo de Protocolos,
# que es la referencia de identidad de toda la suite: los tres módulos tienen
# que nombrar al mismo proyecto igual.
PROYECTOS = {
    "P2416 - MLP - DESALADORA": {"id": "DESALADORA", "codigo": "P2416",
                                 "nombre": "Desaladora",
                                 "cliente": "ANTOFAGASTA MINERALS", "obra": True},
    "P2407 - CODELCO - TALABRE": {"id": "TALABRE", "codigo": "P2407",
                                  "nombre": "Talabre",
                                  "cliente": "CODELCO", "obra": True},
    "P2342 - ARQUEROS MASA": {"id": "ARQUEROS", "codigo": "P2342",
                              "nombre": "Arqueros",
                              "cliente": "MASA", "obra": True},
    "OFICINA CENTRAL": {"id": "OFICINA", "codigo": "—", "nombre": "Oficina Central",
                        "cliente": "BESALCO MONTAJES", "obra": False},
}
ORDEN = ["DESALADORA", "TALABRE", "ARQUEROS", "OFICINA"]

CERRADA, ABIERTA = "Cerrada", "Abierta"
TIPOS = ["No Conformidad", "Producto No Conforme", "Observación",
         "Opción de Mejora", "Reporte HSE Nivel Alto"]

# Tipos que NO entran al módulo. La opción de mejora es una propuesta, no un
# hallazgo que haya que corregir: mezclarla con las NC infla el universo y
# ensucia el % de cierre con algo que nadie está obligado a cerrar. Se cuentan
# aparte y el panel declara cuántas quedaron fuera, para que la exclusión sea
# visible y no un dato que se perdió.  Confirmado por el usuario (03-08-2026).
TIPOS_EXCLUIDOS = ["Opción de Mejora"]
excluidos = Counter()            # por tipo
excluidos_proy = Counter()       # por proyecto, para declarar de dónde salieron
TRAMOS = [("1-30 días", 0), ("31-90 días", 30), ("91-180 días", 90),
          ("181-365 días", 180), ("Más de un año", 365)]

# Ritmo semanal: la ventana sigue siendo de 8 semanas, pero solo las 5 más
# recientes se abren una a una; las 3 anteriores se acumulan en una fila.
# Ocho filas ocupaban media pantalla sin decir más que el bloque.
SEM_VENTANA, SEM_DETALLE = 8, 5

# ---------------------------------------------------------------------------
# PLAZO DE RESPUESTA — la regla de atraso del proyecto
# ---------------------------------------------------------------------------
# Hay 10 días para responder una NC desde que se emite; a partir del día 11
# corre atraso. Se cuenta sobre la FECHA DE EMISIÓN, que las dos fuentes sí
# traen, y no sobre una fecha comprometida, que ninguna trae. Por eso el atraso
# SÍ es calculable, al revés de lo que suponía la versión anterior.
#     atrasada     = abierta con más de PLAZO_RESPUESTA días desde que se emitió
#     días de atraso = antigüedad − PLAZO_RESPUESTA
# Confirmado por el usuario (03-08-2026). Si el plazo cambia, se cambia acá y
# el panel, la PPT y la suite se recalculan solos.
PLAZO_RESPUESTA = 10

# Las tres vías por las que entra un hallazgo. Se distinguen porque no son lo
# mismo de gestionar: la del cliente compromete el contrato, la del subcontrato
# la absorbe Besalco, y la interna es autodetección.
EMISIONES = ["Interna BSMT", "Externa Cliente", "Externa Subcontrato"]
EMISION_CORTA = {"Interna BSMT": "Interna (Besalco)",
                 "Externa Cliente": "Externa · Cliente",
                 "Externa Subcontrato": "Externa · Subcontrato",
                 "Sin clasificar": "Sin clasificar"}
# La misma vía, redactada para meterla en una frase («4 del cliente»).
EMISION_FRASE = {"Interna BSMT": "de Besalco",
                 "Externa Cliente": "del cliente",
                 "Externa Subcontrato": "de un subcontrato",
                 "Sin clasificar": "sin clasificar"}

# Ventanas para «lo levantado recientemente». 7 días es la que se mira semana a
# semana; las otras dan contexto para saber si la semana fue alta o baja.
VENTANAS = [("semana", 7), ("mes", 30), ("trimestre", 90)]

avisos = []


def norm(v):
    if v is None:
        return ""
    s = str(v).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    return " ".join(s.split())


def pct1(a, b):
    """Porcentaje con 1 decimal, redondeando .5 hacia arriba (igual que el panel)."""
    return floor(1000.0 * a / b + 0.5) / 10.0 if b else 0.0


def texto(v, alt=""):
    return str(v).strip() if v not in (None, "") else alt


def semana_cerrada(hoy):
    """Última semana calendario COMPLETA (lunes a domingo) al corte.

    El informe se arma el lunes y habla de la semana que acaba de cerrar, no de
    los últimos 7 días corridos: con un corte el lunes 03-08 la ventana es
    lunes 27-07 → domingo 02-08. Antes se usaba `creada > hoy − 7 días`, que
    daba 28-07 → 03-08: metía el lunes del propio informe —un día que recién
    empieza— y dejaba fuera el lunes anterior, partiendo la semana en dos.

    Si el corte cae un domingo esa semana ya terminó y es la que se informa; si
    cae a mitad de semana, se informa la anterior completa.
    """
    fin = hoy - timedelta(days=(hoy.weekday() + 1) % 7)   # domingo ya cerrado
    return fin - timedelta(days=6), fin                    # lunes, domingo


def dias_atraso(estado, creada, hoy):
    """Días de atraso de una NC abierta: los que pasan del plazo de respuesta.

    Se cuenta desde la emisión, no desde una fecha comprometida: esa no la trae
    ninguna de las dos fuentes, la de emisión sí. En el día 11 el atraso es 1.
    Devuelve None si no aplica (cerrada, sin fecha, o todavía en plazo).
    """
    if estado != ABIERTA or not creada:
        return None
    d = (hoy - creada).days - PLAZO_RESPUESTA
    return d if d > 0 else None


def atrasada(estado, creada, hoy):
    return dias_atraso(estado, creada, hoy) is not None


# =============================================================================
def leer(ruta, hoy):
    wb = load_workbook(ruta, data_only=True)
    hoja = "Observaciones" if "Observaciones" in wb.sheetnames else wb.sheetnames[0]
    it = wb[hoja].iter_rows(values_only=True)
    next(it)                                     # encabezado

    items, sin_proyecto = [], Counter()
    for f in it:
        if all(x is None for x in f):
            continue
        clave = texto(f[0]).upper()
        p = PROYECTOS.get(clave)
        if p is None:
            sin_proyecto[clave or "(vacío)"] += 1
            continue

        emision = texto(f[3], "Sin clasificar")
        if norm(emision).startswith("interna"):
            origen = "Interna"
        elif norm(emision).startswith("externa"):
            origen = "Externa"
        else:
            origen = "Sin clasificar"
            # Vacío es un dato faltante, no un valor raro: se cuenta aparte y
            # solo se avisa si aparece un texto que no se sabe clasificar.
            if f[3] not in (None, ""):
                avisos.append(f"Tipo de emisión no reconocido: «{emision}»")

        estatus = texto(f[24], "Sin estatus")
        estado = CERRADA if norm(estatus) == "cerrado" else ABIERTA

        creada = f[11] if isinstance(f[11], datetime) else None
        cerrada = f[21] if isinstance(f[21], datetime) else None

        tipo = texto(f[5], "Sin tipo")
        if tipo not in TIPOS:
            avisos.append(f"Tipo no reconocido: «{tipo}»")
        if tipo in TIPOS_EXCLUIDOS:
            excluidos[tipo] += 1
            excluidos_proy[p["id"]] += 1
            continue

        items.append({
            "proy": p["id"],
            "n": texto(f[1]),
            "origen": origen,
            "emision": emision,
            "codigoExterno": texto(f[4]),
            "tipo": tipo,
            "especialidad": texto(f[8], "Sin especialidad"),
            "responsable": texto(f[9], "Sin asignar"),
            "creada": creada,
            "titulo": texto(f[12])[:150],
            "costo": f[19] if isinstance(f[19], (int, float)) else None,
            "cerradaEl": cerrada,
            "estatus": estatus,
            "estado": estado,
            # Regla del proyecto: 10 días para responder desde que se emite;
            # del día 11 en adelante corre atraso. Ver PLAZO_RESPUESTA.
            "atrasada": atrasada(estado, creada, hoy),
            "diasAtraso": dias_atraso(estado, creada, hoy),
            "diasAbierta": (hoy - creada).days if estado == ABIERTA and creada else None,
            "diasCierre": (cerrada - creada).days if estado == CERRADA and cerrada and creada else None,
        })

    for k, n in sin_proyecto.items():
        avisos.append(f"{n} registros con proyecto no reconocido «{k}» — quedaron fuera")
    return items


# =============================================================================
# NC EXTERNAS DEL CLIENTE — segunda fuente, solo Arqueros/MASA
# =============================================================================
# El registro principal («Observaciones») trae, para Arqueros, únicamente lo que
# emite Besalco: internas del contrato y contra subcontratos. Las que **el
# cliente MASA levanta contra Besalco** viven en su propia planilla de control
# de recepción y respuesta, y sin ellas el frente se ve casi sin pendiente.
#
# Hoja «Disposición NC-Externas» · encabezado fila 4 · datos fila 5+
#   1  N° NC          2  Tipo (NC/OBS/OdM)   3  Interna/Externa   4  Detalle
#   5  TTAL de Recepción                     6  Fecha de Recepción
#   7  Status Actual
#   (14,13) (21,20) (28,27) = (Status, Fecha MASA) de las revisiones 0, 1 y 2
#
# TRAMPA: la «Fecha MASA» de las revisiones abiertas trae la fecha de HOY —es una
# fórmula que cuenta días de espera—, así que solo se lee como fecha de cierre
# cuando el status de esa misma revisión es «Aprobado».
HOJA_EXTERNAS = "Disposición NC-Externas"
TIPO_EXTERNAS = {"no conformidad": "No Conformidad", "observacion": "Observación",
                 "odm": "Opción de Mejora", "opcion de mejora": "Opción de Mejora"}
# Estatus del cliente. «Cerrada» es el único que cierra; los demás siguen abiertos
# y se conservan tal cual porque no se gestionan igual: «Observada» volvió con
# observaciones del cliente y «Pendiente» aún no tiene respuesta de Besalco.
CIERRE_EXTERNAS = "cerrada"


def leer_externas(ruta, hoy, proyecto="ARQUEROS"):
    """NC que el cliente levanta contra Besalco, desde su planilla de control."""
    wb = load_workbook(ruta, data_only=True)
    if HOJA_EXTERNAS not in wb.sheetnames:
        avisos.append(f"El archivo de externas no trae la hoja «{HOJA_EXTERNAS}»; se ignoró")
        return []
    ws = wb[HOJA_EXTERNAS]

    items = []
    for f in ws.iter_rows(min_row=5, values_only=True):
        if not f or (f[1] in (None, "") and f[4] in (None, "")):
            continue
        if norm(f[3]) and not norm(f[3]).startswith("externa"):
            avisos.append(f"Fila «{texto(f[1])}» del archivo de externas marcada "
                          f"«{texto(f[3])}» y no «Externa» — se omitió")
            continue

        estatus = texto(f[7], "Sin estatus")
        estado = CERRADA if norm(estatus) == CIERRE_EXTERNAS else ABIERTA

        creada = f[6] if isinstance(f[6], datetime) else None
        cerrada = None
        for st, fe in ((14, 13), (21, 20), (28, 27)):
            if norm(f[st]) == "aprobado" and isinstance(f[fe], datetime):
                cerrada = f[fe]
        if estado == ABIERTA:
            cerrada = None          # la fecha de las abiertas es =HOY(), no un cierre

        tipo = TIPO_EXTERNAS.get(norm(f[2]))
        if tipo is None:
            tipo = texto(f[2], "Sin tipo")
            avisos.append(f"Tipo no reconocido en el archivo de externas: «{tipo}»")
        if tipo in TIPOS_EXCLUIDOS:
            excluidos[tipo] += 1
            excluidos_proy[proyecto] += 1
            continue

        items.append({
            "proy": proyecto,
            # Se prefija para que no colisione con la numeración interna del
            # proyecto, que corre por su cuenta.
            "n": f"MASA-{texto(f[1])}" if texto(f[1]) else "MASA-s/n",
            "fuente": "externas",
            "origen": "Externa",
            "emision": "Externa Cliente",
            "codigoExterno": texto(f[5]),
            "tipo": tipo,
            # La planilla del cliente no registra ni disciplina ni responsable:
            # se declara faltante en vez de inventarlo.
            "especialidad": "Sin especialidad",
            "responsable": "Sin asignar",
            "creada": creada,
            "titulo": texto(f[4])[:150],
            "costo": None,
            "cerradaEl": cerrada,
            "estatus": estatus,
            "estado": estado,
            # Misma regla que el registro principal: la planilla del cliente
            # también trae la fecha de recepción, que es cuando se emite.
            "atrasada": atrasada(estado, creada, hoy),
            "diasAtraso": dias_atraso(estado, creada, hoy),
            "diasAbierta": (hoy - creada).days if estado == ABIERTA and creada else None,
            "diasCierre": (cerrada - creada).days if estado == CERRADA and cerrada and creada else None,
        })

    if items:
        ab = sum(1 for i in items if i["estado"] == ABIERTA)
        print(f"  NC externas del cliente ({proyecto}): {len(items)} registros · {ab} abiertas")
    return items


# =============================================================================
def resumir(items):
    ab = [i for i in items if i["estado"] == ABIERTA]
    cer = [i for i in items if i["estado"] == CERRADA]
    dc = [i["diasCierre"] for i in cer if i["diasCierre"] is not None]
    costos = [i["costo"] for i in items if isinstance(i["costo"], (int, float))]
    return {
        "total": len(items),
        "cerradas": len(cer),
        "abiertas": len(ab),
        "atrasadas": sum(1 for i in ab if i["atrasada"]),
        "pctCierre": pct1(len(cer), len(items)),
        "internas": sum(1 for i in items if i["origen"] == "Interna"),
        "externas": sum(1 for i in items if i["origen"] == "Externa"),
        # Las dos externas NO son lo mismo y no se gestionan igual: la del
        # cliente compromete el contrato, la del subcontrato la absorbe
        # Besalco. Se cuentan por separado —también sobre lo abierto— para que
        # el panel nunca tenga que sumarlas en un solo «externas».
        "cliente": sum(1 for i in items if i["emision"] == "Externa Cliente"),
        "subcontrato": sum(1 for i in items if i["emision"] == "Externa Subcontrato"),
        "abiertasCliente": sum(1 for i in ab if i["emision"] == "Externa Cliente"),
        "abiertasSubcontrato": sum(1 for i in ab if i["emision"] == "Externa Subcontrato"),
        "abiertasInternas": sum(1 for i in ab if i["origen"] == "Interna"),
        "costo": round(sum(costos), 1),
        "conCosto": len(costos),
        "medianaCierre": int(median(dc)) if dc else None,
        # Promedio, además de la mediana: el promedio es el que pide el cuadro
        # de «estado de cada vía» y la mediana el que va en los KPI. Se guardan
        # los dos porque no dicen lo mismo — un solo hallazgo de 400 días mueve
        # el promedio y no la mediana.
        "promedioCierre": int(round(sum(dc) / len(dc))) if dc else None,
        "maxCierre": max(dc) if dc else None,
    }


def tramo(d):
    t = TRAMOS[0][0]
    for nom, piso in TRAMOS:
        if d > piso:
            t = nom
    return t


def construir(items, hoy, fuente):
    proys = [p for p in ORDEN if any(i["proy"] == p for i in items)]
    meta_p = {v["id"]: v for v in PROYECTOS.values()}

    tipos = [t for t in TIPOS if t not in TIPOS_EXCLUIDOS
             and any(i["tipo"] == t for i in items)]
    esps = [e for e, _ in Counter(i["especialidad"] for i in items).most_common()]

    def novedades(sub, hoy=hoy):
        """Lo levantado en la última semana cerrada, con su detalle y el ritmo previo.

        Es la lista que se revisa en la reunión semanal: cuántos entraron, en
        qué disciplina y si los detectó Besalco o los levantó el cliente. La
        tendencia de las 8 semanas anteriores va al lado porque el número
        suelto no dice nada: 4 hallazgos son pocos o muchos según el ritmo.

        La ventana es la semana calendario lunes-domingo que ya cerró, y todas
        las semanas del ritmo se cuentan hacia atrás desde ese domingo.
        """
        ini_sem, fin_sem = semana_cerrada(hoy)
        # Las fechas del Excel vienen a medianoche: el domingo entra completo
        # comparando contra el lunes siguiente.
        en = lambda i, a, b_: i["creada"] and a <= i["creada"] < b_ + timedelta(days=1)
        nuevas = sorted([i for i in sub if en(i, ini_sem, fin_sem)],
                        key=lambda x: (x["creada"], x["proy"]))

        def cortar(s_):
            return {
                "total": len(s_),
                "internas": sum(1 for i in s_ if i["origen"] == "Interna"),
                "externas": sum(1 for i in s_ if i["origen"] == "Externa"),
                # Mismo desglose de tres vías que `resumir()`: la externa del
                # cliente se informa aparte de la del subcontrato.
                "cliente": sum(1 for i in s_ if i["emision"] == "Externa Cliente"),
                "subcontrato": sum(1 for i in s_ if i["emision"] == "Externa Subcontrato"),
                "abiertas": sum(1 for i in s_ if i["estado"] == ABIERTA),
            }

        # Las SEM_DETALLE semanas más recientes van una a una; el resto de la
        # ventana se acumula en una sola fila. Con las 8 semanas abiertas la
        # sección crecía de más en vertical sin aportar lectura: las semanas
        # viejas se leen igual como bloque. La fila acumulada se marca para que
        # el panel avise cuántas semanas suma y nadie la compare con una sola.
        # Todas las semanas del ritmo son lunes-domingo, contadas hacia atrás
        # desde el domingo de la semana cerrada.
        sem = lambda k: (ini_sem - timedelta(days=7 * k), fin_sem - timedelta(days=7 * k))
        tendencia = []
        ini, _ = sem(SEM_VENTANA - 1)
        _, fin = sem(SEM_DETALLE)
        tendencia.append({
            "etiqueta": f"{SEM_DETALLE} semanas o más",
            "desde": ini.strftime("%d-%m"),
            "hasta": fin.strftime("%d-%m"),
            "semanas": SEM_VENTANA - SEM_DETALLE,
            "acumulada": True,
            **cortar([i for i in sub if en(i, ini, fin)]),
        })
        for k in range(SEM_DETALLE - 1, -1, -1):
            ini, fin = sem(k)
            tendencia.append({
                # «Última semana» es la que se informa —la que ya cerró—, no la
                # que corre: el informe se arma el lunes siguiente.
                "etiqueta": ("Última semana" if k == 0 else
                             "Semana anterior" if k == 1 else f"Hace {k} semanas"),
                "desde": ini.strftime("%d-%m"),
                "hasta": fin.strftime("%d-%m"),
                "semanas": 1,
                "acumulada": False,
                **cortar([i for i in sub if en(i, ini, fin)]),
            })

        return {
            # Lunes y domingo de la semana informada, ambos incluidos.
            "desde": ini_sem.strftime("%d-%m-%Y"),
            "hasta": fin_sem.strftime("%d-%m-%Y"),
            **cortar(nuevas),
            "cerradas": sum(1 for i in nuevas if i["estado"] == CERRADA),
            "porOrigen": dict(Counter(i["origen"] for i in nuevas)),
            "porEmision": dict(Counter(i["emision"] for i in nuevas)),
            # Los tres cortes que se piden en la reunión —disciplina, frente y
            # tipo— van con el mismo desglose interna/externa, para que el
            # panel los dibuje todos igual.
            "porEspecialidad": {e: cortar([i for i in nuevas if i["especialidad"] == e])
                                for e, _ in Counter(i["especialidad"]
                                                    for i in nuevas).most_common()},
            "porProyecto": {n: cortar([i for i in nuevas
                                       if meta_p[i["proy"]]["nombre"] == n])
                            for n, _ in Counter(meta_p[i["proy"]]["nombre"]
                                                for i in nuevas).most_common()},
            "porTipo": {t: cortar([i for i in nuevas if i["tipo"] == t])
                        for t, _ in Counter(i["tipo"] for i in nuevas).most_common()},
            "detalle": [{
                "fecha": i["creada"].strftime("%d-%m-%Y"),
                "proy": meta_p[i["proy"]]["nombre"], "n": i["n"], "tipo": i["tipo"],
                "emision": i["emision"], "origen": i["origen"],
                "especialidad": i["especialidad"], "responsable": i["responsable"],
                "estatus": i["estatus"], "estado": i["estado"],
                "titulo": i["titulo"], "costo": i["costo"],
                "codigoExterno": i["codigoExterno"],
            } for i in nuevas],
            "tendencia": tendencia,
        }

    def bloque(sub, hoy=hoy):
        ab = [i for i in sub if i["estado"] == ABIERTA]
        _, fin_semana = semana_cerrada(hoy)
        return {
            "resumen": resumir(sub),
            "porTipo": {t: resumir([i for i in sub if i["tipo"] == t]) for t in tipos
                        if any(i["tipo"] == t for i in sub)},
            "porOrigen": {o: resumir([i for i in sub if i["origen"] == o])
                          for o in ("Interna", "Externa", "Sin clasificar")
                          if any(i["origen"] == o for i in sub)},
            # Emisión con su cierre, no solo el conteo: «Externa Cliente» y
            # «Externa Subcontrato» se comportan distinto y hay que verlo.
            "porEmision": {e: resumir([i for i in sub if i["emision"] == e])
                           for e in EMISIONES + ["Sin clasificar"]
                           if any(i["emision"] == e for i in sub)},
            # Autodetección: qué parte de los hallazgos los levanta Besalco.
            "autodeteccion": pct1(sum(1 for i in sub if i["origen"] == "Interna"),
                                  sum(1 for i in sub if i["origen"] in ("Interna", "Externa"))),
            "porEspecialidad": {e: resumir([i for i in sub if i["especialidad"] == e])
                                for e in esps if any(i["especialidad"] == e for i in sub)},
            "porEstatus": dict(Counter(i["estatus"] for i in ab)),
            "antiguedad": {t: sum(1 for i in ab if i["diasAbierta"] is not None
                                  and tramo(i["diasAbierta"]) == t)
                           for t, _ in TRAMOS},
            "porAnio": dict(sorted(Counter(i["creada"].year for i in sub if i["creada"]).items())),
            # Lo levantado en cada ventana reciente, contado sobre el subconjunto.
            # Se cuentan hacia atrás desde el DOMINGO de la semana informada,
            # no desde el corte: si no, «los últimos 30 días» incluirían días
            # que la semana del informe deja fuera.
            "ventanas": {n: resumir([i for i in sub if i["creada"]
                                     and 0 <= (fin_semana - i["creada"]).days < d])
                         for n, d in VENTANAS},
            # Lo levantado en la última semana, con detalle. Va dentro del
            # bloque para que cada pestaña muestre lo suyo: el corporativo los
            # cuatro frentes y cada proyecto solo los propios.
            "semana": novedades(sub),
            "porMes": dict(sorted(Counter(i["creada"].strftime("%Y-%m")
                                          for i in sub if i["creada"]).items())),
        }

    datos = {
        "meta": {
            **MODULO,
            "hoy": hoy.strftime("%d-%m-%Y"),
            "corte": hoy.strftime("%Y-%m-%d"),
            "corteTexto": hoy.strftime("%d-%m-%Y"),
            "generado": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "fuente": fuente,
        },
        "orden": proys,
        "proyectos": {p: {**meta_p[p], **bloque([i for i in items if i["proy"] == p])}
                      for p in proys},
        "global": bloque(items),
        # Los tres proyectos de obra, sin Oficina Central: es lo comparable con
        # los otros módulos de la suite.
        "obra": bloque([i for i in items if meta_p[i["proy"]]["obra"]]),
        "tipos": tipos,
        "especialidades": esps,
        # Las etiquetas viven acá y no en el HTML: si mañana aparece una vía de
        # emisión nueva, se agrega en un solo lado.
        "etiquetasEmision": EMISION_CORTA,
        "frasesEmision": EMISION_FRASE,
    }

    # ---- detalle de las abiertas: es la lista accionable ----
    ab = sorted([i for i in items if i["estado"] == ABIERTA],
                key=lambda x: -(x["diasAbierta"] or 0))
    datos["abiertas"] = [{
        "proy": i["proy"], "proyNombre": meta_p[i["proy"]]["nombre"], "n": i["n"],
        "tipo": i["tipo"], "origen": i["origen"], "emision": i["emision"],
        "especialidad": i["especialidad"], "responsable": i["responsable"],
        "titulo": i["titulo"], "estatus": i["estatus"],
        "creada": i["creada"].strftime("%d-%m-%Y") if i["creada"] else None,
        "dias": i["diasAbierta"], "atrasada": i["atrasada"],
        "diasAtraso": i["diasAtraso"], "costo": i["costo"],
    } for i in ab]

    # ---- responsables con NC abiertas ----
    resp = Counter(i["responsable"] for i in ab)
    datos["responsables"] = [{"nombre": n, "abiertas": c,
                              "dias": max((x["diasAbierta"] or 0) for x in ab
                                          if x["responsable"] == n)}
                             for n, c in resp.most_common(12)]

    # ---- control de calidad del dato ----
    nAb = datos["global"]["resumen"]["abiertas"]
    ab_todas = [i for i in items if i["estado"] == ABIERTA]
    sinFecha = sum(1 for i in ab_todas if not i["creada"])
    datos["control"] = {
        "registros": len(items),
        "abiertas": nAb,
        # El atraso se mide contra el PLAZO DE RESPUESTA desde la emisión, no
        # contra una fecha comprometida —que ninguna fuente trae—. Como la fecha
        # de emisión sí está en las dos, el atraso es calculable para todas.
        "plazoRespuesta": PLAZO_RESPUESTA,
        "atrasoCalculable": nAb == 0 or sinFecha < nAb,
        "abiertasSinFechaEmision": sinFecha,
        "rangoFechas": [min(i["creada"] for i in items if i["creada"]).strftime("%d-%m-%Y"),
                        max(i["creada"] for i in items if i["creada"]).strftime("%d-%m-%Y")],
        # Lo que se dejó fuera a propósito: se declara, no se calla.
        "excluidos": dict(excluidos),
        # De qué frente salió cada uno: la exclusión rige para los cuatro, y
        # verlo desglosado evita la pregunta de si se aplicó en todos.
        "excluidosPorProyecto": {meta_p[k]["nombre"]: n
                                 for k, n in excluidos_proy.most_common()},
        "tiposExcluidos": TIPOS_EXCLUIDOS,
        "sinCosto": sum(1 for i in items if i["costo"] is None),
        "sinEspecialidad": sum(1 for i in items if i["especialidad"] == "Sin especialidad"),
        # Las NC del cliente vienen de su propia planilla, que no registra
        # disciplina, responsable ni costo. Se declara para que no parezca
        # que el dato se perdió.
        "externasCliente": {
            "registros": sum(1 for i in items if i.get("fuente") == "externas"),
            "sinEspecialidad": sum(1 for i in items
                                   if i.get("fuente") == "externas"
                                   and i["especialidad"] == "Sin especialidad"),
        },
    }
    if excluidos:
        avisos.append("Fuera del módulo por tipo: "
                      + " · ".join(f"{n} {t}" for t, n in excluidos.items())
                      + " — de "
                      + " · ".join(f"{meta_p[k]['nombre']} {n}"
                                   for k, n in excluidos_proy.most_common())
                      + " (son propuestas, no hallazgos que corregir; el panel lo declara)")
    if sinFecha:
        avisos.append(
            f"{sinFecha} NC abiertas no traen fecha de emisión: no se les puede calcular "
            f"el atraso y quedan fuera del conteo de atrasadas.")

    # ---- semáforo por proyecto ----
    datos["semaforo"] = {}
    for p in proys:
        r = datos["proyectos"][p]["resumen"]
        if r["abiertas"] > 15 or r["pctCierre"] < 85:
            est = ["crit", "Crítico"]
        elif r["abiertas"] > 3 or r["pctCierre"] < 95:
            est = ["warn", "Atención"]
        else:
            est = ["good", "Al día"]
        datos["semaforo"][p] = est
    return datos


MARCA_INI = "<!-- === NC:INICIO ==="
MARCA_FIN = "<!-- === NC:FIN === -->"


def inyectar_en_html(ruta_html, datos):
    """Reemplaza el bloque de datos dentro de index.html."""
    if not ruta_html.exists():
        avisos.append(f"No se encontró {ruta_html.name}; solo se escribió el JSON.")
        return False
    html = ruta_html.read_text(encoding="utf-8")
    i, f = html.find(MARCA_INI), html.find(MARCA_FIN)
    if i == -1 or f == -1 or f < i:
        avisos.append(f"No se encontraron las marcas NC en {ruta_html.name}.")
        return False
    bloque = (MARCA_INI + " (bloque generado por no_conformidades.py — no editar a mano) -->\n"
              "<script>\nconst NC = " + json.dumps(datos, ensure_ascii=False, indent=1)
              + ";\n</script>\n")
    ruta_html.write_text(html[:i] + bloque + html[f:], encoding="utf-8")
    return True


# =============================================================================
def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--data", required=True, help="Data_NCR*.xlsx")
    ap.add_argument("--externas", help="Data_NCR*externas.xlsx — NC que el cliente MASA "
                                       "levanta en Arqueros; el registro principal no las trae")
    ap.add_argument("--hoy", help="Fecha de referencia AAAA-MM-DD (por defecto, hoy)")
    args = ap.parse_args()

    hoy = datetime.strptime(args.hoy, "%Y-%m-%d") if args.hoy else \
        datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    items = leer(args.data, hoy)
    if not items:
        sys.exit("No se leyó ningún registro. Revisa el archivo.")
    n_ext = 0
    if args.externas:
        ext = leer_externas(args.externas, hoy)
        n_ext = len(ext)
        items += ext
    else:
        avisos.append("No se pasó --externas: Arqueros queda sin las NC que le levanta "
                      "el cliente MASA, que no vienen en el registro principal.")
    datos = construir(items, hoy, Path(args.data).name)

    destino = AQUI / "datos_nc.json"
    destino.write_text(json.dumps(datos, ensure_ascii=False, indent=1), encoding="utf-8")
    inyectado = inyectar_en_html(AQUI / "index.html", datos)

    # ---------------- resumen en pantalla ----------------
    pc = lambda a, b: f"{pct1(a, b):.1f}%".replace(".", ",") if b else "—"
    g = datos["global"]["resumen"]
    print("=" * 70)
    print(f"  NO CONFORMIDADES — {datos['meta']['fuente']}")
    print(f"  {g['total']} registros · referencia {datos['meta']['hoy']}")
    print("=" * 70)

    print(f"\n{'PROYECTO':17s} {'total':>6s} {'cerr':>5s} {'abier':>6s} {'%cierre':>8s} "
          f"{'inter':>6s} {'exter':>6s} {'UF':>8s} {'med.cierre':>11s}")
    for p in datos["orden"]:
        r = datos["proyectos"][p]["resumen"]
        n = datos["proyectos"][p]["nombre"]
        mc = f"{r['medianaCierre']} d" if r["medianaCierre"] is not None else "—"
        print(f"{n:17s} {r['total']:6d} {r['cerradas']:5d} {r['abiertas']:6d} "
              f"{pc(r['cerradas'], r['total']):>8s} {r['internas']:6d} {r['externas']:6d} "
              f"{r['costo']:8,.0f}".replace(",", ".") + f" {mc:>11s}")
    print(f"{'TOTAL':17s} {g['total']:6d} {g['cerradas']:5d} {g['abiertas']:6d} "
          f"{pc(g['cerradas'], g['total']):>8s} {g['internas']:6d} {g['externas']:6d} "
          f"{g['costo']:8,.0f}".replace(",", ".") +
          f" {(str(g['medianaCierre']) + ' d') if g['medianaCierre'] is not None else '—':>11s}")

    S = datos["global"]["semana"]
    print(f"\nLEVANTADOS EN LA ÚLTIMA SEMANA  ({S['desde']} → {S['hasta']}): {S['total']}")
    if S["total"]:
        print(f"  {S['internas']} internas · {S['externas']} externas · "
              f"{S['abiertas']} siguen abiertas")
        for e, r in S["porEspecialidad"].items():
            print(f"    {e:18s} {r['total']:3d}  ({r['internas']} int · {r['externas']} ext)")
        for d in S["detalle"]:
            print(f"    {d['fecha']}  {d['proy']:11s} N°{str(d['n'] or '—'):>5s}  "
                  f"{EMISION_CORTA.get(d['emision'], d['emision']):22s} "
                  f"{d['especialidad']:16s} {d['estatus']}")
    print("  ritmo de las últimas 8 semanas: " +
          " ".join(str(t["total"]) for t in S["tendencia"]))

    print("\nPOR TIPO")
    for t, r in datos["global"]["porTipo"].items():
        print(f"  {t:24s} {r['total']:4d} · abiertas {r['abiertas']:2d}")

    print("\nESPECIALIDADES con NC abiertas")
    for e, r in sorted(datos["global"]["porEspecialidad"].items(),
                       key=lambda x: -x[1]["abiertas"]):
        if r["abiertas"]:
            print(f"  {e:18s} {r['total']:4d} levantadas · {r['abiertas']:2d} abiertas")

    print("\nANTIGÜEDAD de las NC abiertas")
    for t, n in datos["global"]["antiguedad"].items():
        if n:
            print(f"  {t:15s} {n:3d}")

    K = datos["control"]
    print("\nCONTROL DE CALIDAD DEL DATO")
    print(f"  Registros procesados: {K['registros']} · rango {K['rangoFechas'][0]} → {K['rangoFechas'][1]}")
    print(f"  NC abiertas: {K['abiertas']} · plazo de respuesta: {K['plazoRespuesta']} días "
          f"desde la emisión")
    if K["atrasoCalculable"]:
        print(f"  OK  Atraso calculable: {g['atrasadas']} NC atrasadas de {K['abiertas']} abiertas")
    else:
        print("  ⚠  El ATRASO NO ES CALCULABLE: las abiertas no traen fecha de emisión")
    print(f"  Sin costo declarado: {K['sinCosto']} · sin especialidad: {K['sinEspecialidad']}")

    if avisos:
        print("\n⚠  AVISOS — revisar antes de publicar:")
        for a in sorted(set(avisos))[:15]:
            print(f"   · {a}")

    print(f"\nEscritos: {destino.name}" + (", index.html" if inyectado else ""))


if __name__ == "__main__":
    main()
