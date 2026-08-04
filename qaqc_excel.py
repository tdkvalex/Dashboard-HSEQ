#!/usr/bin/env python3
"""Apertura de los Excel de origen: rápida, y con errores que se entienden.

Lo usan los cuatro scripts de módulo y `actualizar_semana.py`. Resuelve tres
cosas que costaban tiempo o confusión cada lunes:

1. **Los nombres de hoja salen del ZIP, sin openpyxl.** Es lo único que hace
   falta para reconocer casi todos los archivos, y abrir el libro entero para
   eso costaba hasta 8 segundos por archivo.

2. **Se descartan los estilos con nombre antes de abrir.** El REPORTE_GERENCIAL
   arrastra 52.000 estilos heredados de años de copiar y pegar entre planillas:
   11,8 MB de `styles.xml` de los que solo 0,13 MB —el bloque `cellXfs`— dice
   algo sobre los datos. openpyxl los parsea todos igual. Quitando `cellStyles`
   y `cellStyleXfs` el archivo abre en 0,2 s en vez de 13 s, y los formatos de
   número quedan intactos: las fechas se siguen leyendo como fechas, que es lo
   único que aquí depende del estilo.

3. **Los fallos se explican en una línea.** Falta el archivo, vino vacío, no es
   un Excel o no trae la hoja esperada: eso se dice, no se vuelca un traceback.

Ninguna lectura de la suite usa `max_row`, `.cell()` ni celdas combinadas —todo
es `iter_rows(values_only=True)`—, así que los libros se abren en modo
`read_only`, que evita construir un objeto por celda.
"""

import io
import re
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    sys.exit("Falta openpyxl.  Instálalo con:  pip install openpyxl")

# Por debajo de esto no vale la pena reescribir el ZIP: se abre tal cual.
UMBRAL_ESTILOS = 512 * 1024

_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_CELL_STYLE_XFS = ('<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" '
                   'fillId="0" borderId="0"/></cellStyleXfs>')
_CELL_STYLES = ('<cellStyles count="1"><cellStyle name="Normal" xfId="0" '
                'builtinId="0"/></cellStyles>')


def _revisar(p):
    """Motivos por los que un archivo no se puede ni intentar abrir."""
    if not p.exists():
        sys.exit(f"No existe el archivo: {p}")
    if p.stat().st_size == 0:
        sys.exit(f"«{p.name}» está vacío (0 bytes): la descarga quedó a medias, "
                 f"bájalo de nuevo")


def hojas(ruta):
    """Nombres de hoja leídos del ZIP. Devuelve [] si no es un Excel legible.

    No usa openpyxl a propósito: reconocer un archivo por sus hojas es lo
    primero que se hace con TODOS los de la carpeta, y abrir cada libro entero
    para eso era el grueso del tiempo de arranque.
    """
    try:
        with zipfile.ZipFile(ruta) as z:
            raiz = ElementTree.fromstring(z.read("xl/workbook.xml"))
        return [h.get("name") for h in raiz.iter(f"{_NS}sheet") if h.get("name")]
    except Exception:
        return []


def _sin_estilos_con_nombre(ruta):
    """Copia en memoria del libro sin los estilos con nombre, o None si no hace
    falta tocarlo. Se conservan `numFmts` y `cellXfs`, que son los que deciden
    qué celda es una fecha; se van `cellStyles` y `cellStyleXfs`, que solo
    nombran formatos para el menú de Excel."""
    try:
        with zipfile.ZipFile(ruta) as z:
            try:
                info = z.getinfo("xl/styles.xml")
            except KeyError:
                return None
            if info.file_size < UMBRAL_ESTILOS:
                return None
            s = z.read("xl/styles.xml").decode("utf-8")
            nuevo = re.sub(r"<cellStyleXfs.*?</cellStyleXfs>", _CELL_STYLE_XFS, s, flags=re.S)
            nuevo = re.sub(r"<cellStyles.*?</cellStyles>", _CELL_STYLES, nuevo, flags=re.S)
            if nuevo == s:
                return None          # el peso no venía de ahí: no vale la pena
            # Los `xfId` de cellXfs apuntaban al bloque que acabamos de reducir
            # a una sola entrada; si no se reapuntan, quedan fuera de rango.
            nuevo = re.sub(r"<cellXfs.*?</cellXfs>",
                           lambda m: re.sub(r'xfId="\d+"', 'xfId="0"', m.group(0)),
                           nuevo, flags=re.S)
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as salida:
                for i in z.infolist():
                    salida.writestr(i, nuevo.encode("utf-8")
                                    if i.filename == "xl/styles.xml" else z.read(i.filename))
            buf.seek(0)
            return buf
    except Exception:
        return None                  # ante la duda, que lo abra openpyxl normal


def libro(ruta, read_only=True, **kw):
    """Abre el Excel. Igual que `load_workbook`, pero rápido y explicándose."""
    p = Path(ruta)
    _revisar(p)
    kw.setdefault("data_only", True)
    origen = _sin_estilos_con_nombre(p) or p
    try:
        return load_workbook(origen, read_only=read_only, **kw)
    except Exception as e:
        if origen is not p:          # el recorte pudo no gustarle: reintenta entero
            try:
                return load_workbook(p, read_only=read_only, **kw)
            except Exception:
                pass
        sys.exit(f"No se pudo abrir «{p.name}» como Excel ({type(e).__name__}). "
                 f"Comprueba que sea .xlsx o .xlsm y que la descarga esté completa.")


def hoja(wb, nombre, ruta):
    """Igual que wb[nombre], pero si la hoja no está dice cuáles sí están."""
    if nombre not in wb.sheetnames:
        sys.exit(f"«{Path(ruta).name}» no trae la hoja «{nombre}». "
                 f"Trae: {', '.join(wb.sheetnames[:8])}"
                 + (" …" if len(wb.sheetnames) > 8 else ""))
    return wb[nombre]
