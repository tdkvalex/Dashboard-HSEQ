#!/usr/bin/env python3
"""
Procesa el proyecto DESALADORA (PV3 Ampliación Desaladora — AMSA/MLP) y escribe
datos_desaladora.json, que alimenta la pestaña Desaladora del panel y la PPT.

Uso (cada semana, con los dos archivos del corte):

    python3 desaladora.py \
        --reporte /ruta/REPORTE_GERENCIAL_AAAAMMDD.xlsx \
        --punch   /ruta/Listado_Puntos_Punch_Consolidado_AAAA_MM_DD.xlsx

Opcional:
    --hoy 2026-07-27     fecha de referencia para calcular atrasos
                         (por defecto, el día en que se corre el script)

------------------------------------------------------------------------------
DE DÓNDE SALE CADA DATO
------------------------------------------------------------------------------
REPORTE GERENCIAL · hoja «REPORTE GERENCIAL» (encabezado fila 10, datos fila 11+)
    col B  Subsistema/Facility        col D  Tipo (Operable/Facility/Componente)
    col E  Zona (Desaladora / EB-2)   col G  Caminata 1   col I  Caminata 2
    col K  Caminata 3                 col AG Estatus del certificado (carpeta)

PUNCH LIST · hoja «LISTADO PUNCH ITEMS» (encabezado fila 8, datos fila 9+)
    col B  Nº Caminata      col G  Subsistema      col H  Disciplina
    col J  Categoría        col P  Fecha Requerida Cierre     col W  STATUS

------------------------------------------------------------------------------
REGLAS DE CLASIFICACIÓN
------------------------------------------------------------------------------
CATEGORÍA  — se homologa al dígito para poder comparar entre proyectos:
    P1A -> P1     P2B -> P2     P3C -> P3     P0 -> P0

ESTADO del punch — tres estados excluyentes:
    Cerrado    STATUS = CERRADO
    Atrasado   STATUS = ABIERTO  y  Fecha Requerida Cierre < hoy
    En plazo   STATUS = ABIERTO  y  Fecha Requerida Cierre >= hoy
    (los abiertos sin fecha requerida se informan aparte, no se dan por atrasados)

CAMINATA — el archivo trae el número con y sin ceros ('001' y '1'): se unifica
    al entero. En las caminatas del reporte gerencial, «Ejecutada» = realizada,
    «N/A» = no aplica a ese tipo de subsistema, vacío o «…...» = pendiente.

UNIVERSO — los subsistemas del proyecto son los Operables y los Facility (95 al
    corte: 80 + 15, tal como los declara la hoja «Resumen general»). Los
    «Componente» son entregas parciales de un subsistema que ya está contado, así
    que NO suman al universo: se informan aparte, con su estado. Ver
    FUERA_DEL_UNIVERSO.

CARPETA — estatus del Certificado de Entrega de Término de Construcción:
    Aprobada CC / Aprobada -> Aprobada        En Rev. AMSA -> En revisión
    Rechazada -> Rechazada                    Proceso BSMT -> En preparación
    vacío -> Sin entregar
    Se considera ENTREGADA al cliente desde que entra a revisión (o sea todas
    menos «En preparación» y «Sin entregar»).
"""

import argparse
import json
import re
import sys
import unicodedata
import warnings
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

warnings.filterwarnings("ignore")

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Falta openpyxl.  Instálalo con:  pip install openpyxl")


def libro(ruta, **kw):
    """Abre el Excel diciendo QUÉ pasa si no se puede, en vez de un traceback.
    El lunes hace falta saber qué archivo falta o llegó a medias, no en qué
    línea de openpyxl se cayó."""
    p = Path(ruta)
    if not p.exists():
        sys.exit(f"No existe el archivo: {p}")
    if p.stat().st_size == 0:
        sys.exit(f"«{p.name}» está vacío (0 bytes): la descarga quedó a medias, bájalo de nuevo")
    try:
        return load_workbook(p, **kw)
    except Exception as e:
        sys.exit(f"No se pudo abrir «{p.name}» como Excel ({type(e).__name__}). "
                 f"Comprueba que sea .xlsx o .xlsm y que la descarga esté completa.")


def hoja(wb, nombre, ruta):
    """Igual que wb[nombre], pero si la hoja no está dice cuáles sí están."""
    if nombre not in wb.sheetnames:
        sys.exit(f"«{Path(ruta).name}» no trae la hoja «{nombre}». "
                 f"Trae: {', '.join(wb.sheetnames[:8])}"
                 + (" …" if len(wb.sheetnames) > 8 else ""))
    return wb[nombre]

AQUI = Path(__file__).resolve().parent

PROYECTO = {
    "id": "DESALADORA",
    "nombre": "Desaladora",
    "descripcion": "PV3 Montaje Electromecánico — Ampliación Desaladora",
    # Código y cliente tal como los escribe el módulo de Protocolos, que es la
    # referencia de identidad de la suite. El mandante directo es Minera Los
    # Pelambres, del grupo Antofagasta Minerals.
    "codigo": "P2416",
    "cliente": "ANTOFAGASTA MINERALS",
    "contrato": "VPP-1001-03-CC-088",
}

# Orden fijo de disciplinas para gráficos y heatmap
DISCIPLINAS = ["Eléctrica", "Piping", "Obras Civiles",
               "Mecánica", "Estructura", "Instrumentación y Control"]
DISC_CORTA = {"Instrumentación y Control": "Instr. y Control",
              "Obras Civiles": "Obras Civiles"}

DISC_MAP = {
    "electrica": "Eléctrica",
    "piping": "Piping",
    "canerias": "Piping",
    "oo.cc": "Obras Civiles",
    "oocc": "Obras Civiles",
    "obras civiles": "Obras Civiles",
    "civil": "Obras Civiles",
    "mecanica": "Mecánica",
    "estructura": "Estructura",
    "instrumentacion": "Instrumentación y Control",
    "instrumentacion y control": "Instrumentación y Control",
}

CARPETA_MAP = {
    "": "Sin entregar",
    "aprobada": "Aprobada",
    "aprobada cc": "Aprobada",
    "en rev. amsa": "En revisión",
    "en rev amsa": "En revisión",
    "en revision amsa": "En revisión",
    "rechazada": "Rechazada",
    "proceso bsmt": "En preparación",
}
# Una carpeta cuenta como ENTREGADA al cliente desde que entra a revisión.
CARPETA_ENTREGADAS = {"En revisión", "Aprobada", "Rechazada"}
CARPETA_ORDEN = ["Aprobada", "En revisión", "Rechazada",
                 "En preparación", "Sin entregar"]

ZONA_NOMBRES = {"Desaladora": "Planta Desaladora", "EB-2": "Estación de Bombeo 2"}

# Tipos que NO son subsistemas del universo. Un «Componente» es una entrega
# parcial de un subsistema que ya está en la lista —«0587-ESL-201 Comp 1» cuelga
# de «0587-ESL-201»—, así que contarlo suma dos veces el mismo alcance físico e
# infla el universo. Quedan fuera del conteo y se informan aparte, con su estado.
# Regla validada por el jefe de calidad del proyecto (30-07-2026) y respaldada
# por la hoja «Resumen general» del propio archivo, que declara 80 operables y
# 15 facility: el universo es 95, no 97.
FUERA_DEL_UNIVERSO = {"Componente"}

avisos = []


def norm(v):
    """Minúsculas, sin tildes, sin espacios sobrantes."""
    if v is None:
        return ""
    s = str(v).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    return " ".join(s.split())


def disciplina(v):
    d = DISC_MAP.get(norm(v))
    if d:
        return d
    if v not in (None, ""):
        avisos.append(f"Disciplina no reconocida: «{v}» — se agrupa en «Otras»")
    return "Otras"


def categoria(v):
    """P1A->P1, P2B->P2, P3C->P3, P0->P0."""
    s = norm(v).upper().replace(" ", "")
    m = re.match(r"^P(\d)", s)
    if m:
        return "P" + m.group(1)
    if s:
        avisos.append(f"Categoría no reconocida: «{v}» — se agrupa en «Sin categoría»")
    return "Sin categoría"


def num_caminata(v):
    """'001' y '1' son la misma caminata."""
    s = re.sub(r"\D", "", str(v or ""))
    return int(s) if s and int(s) in (1, 2, 3) else 0


def pct1(a, b):
    """Porcentaje con 1 decimal, redondeando .5 hacia arriba.

    Se evita round() de Python, que usa redondeo bancario (round(51.25,1) da
    51.2): el panel calcula en JavaScript con Math.round y daría 51,3. Así el
    panel, el JSON y la PPT muestran siempre la misma cifra.
    """
    if not b:
        return 0.0
    from math import floor
    return floor(1000.0 * a / b + 0.5) / 10.0


def as_fecha(v):
    if isinstance(v, datetime):
        return v
    if isinstance(v, str) and v.strip():
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(v.strip(), fmt)
            except ValueError:
                pass
    return None


# =============================================================================
# 1) REPORTE GERENCIAL — subsistemas, caminatas y carpetas
# =============================================================================
def leer_reporte(ruta):
    wb = libro(ruta, data_only=True)
    ws = hoja(wb, "REPORTE GERENCIAL", ruta)

    corte = None
    for f in ws.iter_rows(min_row=1, max_row=9, values_only=True):
        for i, c in enumerate(f):
            if isinstance(c, str) and norm(c) == "fecha":
                for sig in f[i + 1:]:
                    if isinstance(sig, datetime):
                        corte = sig
                        break
    if corte is None:
        avisos.append("No se encontró la fecha de corte en el REPORTE GERENCIAL.")

    def estado_cam(v):
        s = norm(v)
        if s == "ejecutada":
            return "Realizada"
        if s in ("n/a", "na", "no aplica"):
            return "No aplica"
        if s == "" or s.startswith("…") or set(s) <= {".", "…", " "}:
            return "Pendiente"
        avisos.append(f"Estado de caminata no reconocido: «{v}» — se cuenta como pendiente")
        return "Pendiente"

    subs = []
    for f in ws.iter_rows(min_row=11, max_row=181, values_only=True):
        if f[2] in (None, ""):
            continue
        est = norm(f[32])
        carp = CARPETA_MAP.get(est)
        if carp is None:
            carp = "Sin entregar"
            avisos.append(f"Estatus de carpeta no reconocido: «{f[32]}» — se cuenta como sin entregar")
        subs.append({
            "id": str(f[2]).strip(),
            "area": str(f[1]).strip(),
            "tipo": (str(f[3]).strip() if f[3] else "Sin tipo"),
            "zona": (str(f[4]).strip() if f[4] else "Sin zona"),
            "desc": (str(f[5]).strip() if f[5] else ""),
            "cam": {1: estado_cam(f[6]), 2: estado_cam(f[8]), 3: estado_cam(f[10])},
            "carpeta": carp,
            "tarjetaVerde": norm(f[34]) == "si",
        })

    # El propio archivo declara sus totales en «Resumen  general»: se leen para
    # contrastarlos con lo que calculamos y avisar si algo dejó de cuadrar.
    declarado = {}
    if "Resumen  general" in wb.sheetnames:
        rg = list(wb["Resumen  general"].iter_rows(min_row=1, max_row=45, values_only=True))
        etiqueta = None
        for fila in rg:
            txt = " ".join(norm(c) for c in fila if isinstance(c, str))
            if "caminatas 2 subsistemas" in txt:
                etiqueta = "Operable"
            elif "caminatas facility" in txt:
                etiqueta = "Facility"
            elif etiqueta and sum(1 for c in fila if isinstance(c, int)) >= 3:
                nums = [c for c in fila if isinstance(c, int)]
                declarado[etiqueta] = {"total": nums[0], "ejecutadas": nums[1]}
                etiqueta = None

    # El reporte gerencial también lleva su propio conteo de punch por subsistema
    # (columnas N a AC). Se totaliza para contrastarlo con el punch list, que es
    # la fuente ítem a ítem.
    filas = [f for f in ws.iter_rows(min_row=11, max_row=181, values_only=True)
             if f[2] not in (None, "")]
    punch_rep = {}
    for base in (13, 21):                      # Caminata 1-2 y Caminata 3
        for k, cat in enumerate(["P1", "P2", "P3", "P0"]):
            a = sum(f[base + 2 * k] for f in filas if isinstance(f[base + 2 * k], (int, float)))
            c = sum(f[base + 2 * k + 1] for f in filas if isinstance(f[base + 2 * k + 1], (int, float)))
            e = punch_rep.setdefault(cat, {"abiertos": 0, "cerrados": 0})
            e["abiertos"] += int(a)
            e["cerrados"] += int(c)
    declarado["_punch"] = punch_rep
    return corte, subs, declarado


# =============================================================================
# 2) PUNCH LIST — detalles de terminación
# =============================================================================
def leer_punch(ruta, hoy):
    wb = libro(ruta, data_only=True, read_only=True)
    ws = hoja(wb, "LISTADO PUNCH ITEMS", ruta)

    items = []
    for f in ws.iter_rows(min_row=9, values_only=True):
        if f[2] in (None, ""):
            continue
        st = norm(f[22])
        if st == "cerrado":
            estado, atrasado = "Cerrado", False
        elif st == "abierto":
            estado, atrasado = "Abierto", False
        else:
            estado, atrasado = "Abierto", False
            avisos.append(f"STATUS de punch no reconocido: «{f[22]}» — se cuenta como abierto")

        freq = as_fecha(f[15])
        dias = None
        if estado == "Abierto" and freq:
            # Regla del proyecto: abierto con fecha requerida ya vencida = atrasado.
            if freq < hoy:
                atrasado = True
                dias = (hoy - freq).days

        items.append({
            "id": str(f[2]).strip(),
            "cam": num_caminata(f[1]),
            "area": str(f[5]).strip() if f[5] else "Sin área",
            "sub": str(f[6]).strip() if f[6] else "Sin subsistema",
            "disc": disciplina(f[7]),
            "cat": categoria(f[9]),
            "estado": estado,
            "atrasado": atrasado,
            "sinFecha": estado == "Abierto" and freq is None,
            "diasAtraso": dias,
            "resp": str(f[14]).strip() if f[14] else "",
        })
    return items


# =============================================================================
# 3) Agregaciones
# =============================================================================
def resumir(items):
    ab = [i for i in items if i["estado"] == "Abierto"]
    atr = [i for i in ab if i["atrasado"]]
    return {
        "total": len(items),
        "cerrados": sum(1 for i in items if i["estado"] == "Cerrado"),
        "abiertos": len(ab),
        "atrasados": len(atr),
        "enPlazo": len(ab) - len(atr) - sum(1 for i in ab if i["sinFecha"]),
        "sinFecha": sum(1 for i in ab if i["sinFecha"]),
    }


def padre_de(id_comp):
    """Subsistema del que cuelga un componente: «4515-ESL-201-Comp 1» -> «4515-ESL-201»."""
    return re.sub(r"[\s\-_]*comp\.?\s*\d*\s*$", "", id_comp, flags=re.I).strip()


def construir(corte, todos, items, hoy, declarado):
    # ---------- universo del proyecto ----------
    # Los componentes salen del universo (ver FUERA_DEL_UNIVERSO) pero no se
    # pierden: van a su propio bloque con el estado de cada uno.
    subs = [s for s in todos if s["tipo"] not in FUERA_DEL_UNIVERSO]
    fuera = [s for s in todos if s["tipo"] in FUERA_DEL_UNIVERSO]

    ids_univ = {s["id"] for s in subs}
    componentes = {
        "total": len(fuera),
        "entregados": sum(1 for s in fuera if s["carpeta"] in CARPETA_ENTREGADAS),
        "detalle": [{
            "id": s["id"], "tipo": s["tipo"], "zona": s["zona"],
            "padre": padre_de(s["id"]) if padre_de(s["id"]) in ids_univ else None,
            "carpeta": s["carpeta"], "tarjetaVerde": s["tarjetaVerde"],
            "caminatas": [n for n in (1, 2, 3) if s["cam"][n] == "Realizada"],
        } for s in sorted(fuera, key=lambda x: x["id"])],
    }
    for c in componentes["detalle"]:
        if c["padre"] is None:
            avisos.append(
                f"Componente «{c['id']}» no tiene un subsistema padre en el reporte: "
                f"queda fuera del universo igual, pero conviene revisarlo con el proyecto")
        if c["carpeta"] not in CARPETA_ENTREGADAS:
            avisos.append(
                f"Componente «{c['id']}» está fuera del universo pero su carpeta sigue "
                f"«{c['carpeta']}»: no está entregado — confirmar el tratamiento")

    # ---------- subsistemas ----------
    tipos = Counter(s["tipo"] for s in subs)
    zonas = Counter(s["zona"] for s in subs)

    # ---------- caminatas ----------
    caminatas = {}
    for n in (1, 2, 3):
        aplica = [s for s in subs if s["cam"][n] != "No aplica"]
        real = sum(1 for s in aplica if s["cam"][n] == "Realizada")
        caminatas[str(n)] = {
            "aplica": len(aplica),
            "realizada": real,
            "pendiente": len(aplica) - real,
            "noAplica": len(subs) - len(aplica),
            "pct": pct1(real, len(aplica)),
        }

    # Caminatas por tipo de subsistema: el proyecto NO mide lo mismo en cada tipo
    # (Operables van por Caminata 2; Facility por Caminata 1), así que se informa
    # separado en vez de promediar cosas distintas.
    camPorTipo = {}
    for t in tipos:
        st = [s for s in subs if s["tipo"] == t]
        camPorTipo[t] = {"subs": len(st)}
        for n in (1, 2, 3):
            ap = [s for s in st if s["cam"][n] != "No aplica"]
            r = sum(1 for s in ap if s["cam"][n] == "Realizada")
            camPorTipo[t][str(n)] = {
                "aplica": len(ap), "realizada": r, "pendiente": len(ap) - r,
                "pct": pct1(r, len(ap)) if ap else None,
            }

    # Caminata que el proyecto usa como indicador de cada tipo de subsistema.
    # No se infiere: el archivo lo declara en «Resumen general» — los Operables
    # se miden por la Caminata 2 y los Facility por la Caminata 1 — y más abajo
    # se contrasta el cálculo contra esos totales declarados.
    VIGENTE = {"Operable": "2", "Componente": "2", "Facility": "1"}
    for t, v in camPorTipo.items():
        vig = VIGENTE.get(t)
        if vig is None:                       # tipo nuevo: se usa la mayor que aplique
            vig = next((n for n in ("3", "2", "1") if v[n]["aplica"]), None)
            avisos.append(f"Tipo de subsistema nuevo «{t}» — se asumió la caminata {vig} "
                          f"como indicador; confirmar con el proyecto")
        v["vigente"] = vig

    camPorZona = {}
    for z in zonas:
        sz = [s for s in subs if s["zona"] == z]
        camPorZona[z] = {"subs": len(sz)}
        for n in (1, 2, 3):
            ap = [s for s in sz if s["cam"][n] != "No aplica"]
            r = sum(1 for s in ap if s["cam"][n] == "Realizada")
            camPorZona[z][str(n)] = {"aplica": len(ap), "realizada": r,
                                     "pendiente": len(ap) - r}

    pendientes_cam = []
    for s in subs:
        vig = camPorTipo[s["tipo"]]["vigente"]
        if vig and s["cam"][int(vig)] == "Pendiente":
            pendientes_cam.append({"id": s["id"], "tipo": s["tipo"], "zona": s["zona"],
                                   "desc": s["desc"], "caminata": int(vig)})

    # ---------- carpetas ----------
    cnt = Counter(s["carpeta"] for s in subs)
    carpetas = {
        "total": len(subs),
        "estados": {k: cnt.get(k, 0) for k in CARPETA_ORDEN},
        "entregadas": sum(cnt.get(k, 0) for k in CARPETA_ENTREGADAS),
        "aprobadas": cnt.get("Aprobada", 0),
        "rechazadas": cnt.get("Rechazada", 0),
        "enRevision": cnt.get("En revisión", 0),
        "enPreparacion": cnt.get("En preparación", 0),
        "sinEntregar": cnt.get("Sin entregar", 0),
        "tarjetaVerde": sum(1 for s in subs if s["tarjetaVerde"]),
    }
    carpetas["porZona"] = {
        z: {k: sum(1 for s in subs if s["zona"] == z and s["carpeta"] == k)
            for k in CARPETA_ORDEN} for z in zonas}
    carpetas["detalle"] = sorted(
        [{"id": s["id"], "zona": s["zona"], "tipo": s["tipo"], "estado": s["carpeta"]}
         for s in subs if s["carpeta"] != "Sin entregar"],
        key=lambda x: (CARPETA_ORDEN.index(x["estado"]), x["id"]))

    # ---------- punch ----------
    cats = ["P0", "P1", "P2", "P3"]
    if any(i["cat"] == "Sin categoría" for i in items):
        cats.append("Sin categoría")
        avisos.append(f"{sum(1 for i in items if i['cat'] == 'Sin categoría')} ítem(s) de punch "
                      f"sin categoría asignada — no entran en el conteo de P1")
    p1 = [i for i in items if i["cat"] == "P1"]
    areas = sorted({i["area"] for i in items})

    punch = {
        "global": resumir(items),
        "porCategoria": {c: resumir([i for i in items if i["cat"] == c]) for c in cats},
        "porDisciplina": {d: resumir([i for i in items if i["disc"] == d])
                          for d in DISCIPLINAS if any(i["disc"] == d for i in items)},
        "p1PorDisciplina": {d: resumir([i for i in p1 if i["disc"] == d])
                            for d in DISCIPLINAS if any(i["disc"] == d for i in p1)},
        "p1PorArea": {a: resumir([i for i in p1 if i["area"] == a]) for a in areas},
        "porCaminata": {},
        "categorias": cats,
    }
    for n in (1, 2, 3):
        sub = [i for i in items if i["cam"] == n]
        if sub:
            punch["porCaminata"][str(n)] = resumir(sub)
    sin_cam = [i for i in items if i["cam"] == 0]
    if sin_cam:
        punch["porCaminata"]["Sin asignar"] = resumir(sin_cam)

    # heatmap: abiertos P1 por disciplina × caminata
    punch["heatmap"] = {
        "disc": [d for d in DISCIPLINAS if any(i["disc"] == d for i in p1)],
        "cam": [str(n) for n in (1, 2, 3) if any(i["cam"] == n for i in p1)],
    }
    punch["heatmap"]["vals"] = [
        [sum(1 for i in p1 if i["disc"] == d and i["cam"] == int(c)
             and i["estado"] == "Abierto")
         for c in punch["heatmap"]["cam"]]
        for d in punch["heatmap"]["disc"]]

    # antigüedad de los atrasados
    dias = sorted(i["diasAtraso"] for i in items if i["atrasado"] and i["diasAtraso"])
    tramos = Counter(">180 días" if d > 180 else "91-180 días" if d > 90
                     else "31-90 días" if d > 30 else "1-30 días" for d in dias)
    punch["antiguedad"] = {
        "tramos": {k: tramos.get(k, 0) for k in
                   ["1-30 días", "31-90 días", "91-180 días", ">180 días"]},
        "max": dias[-1] if dias else 0,
        "mediana": dias[len(dias) // 2] if dias else 0,
    }

    # subsistemas con más P1 abiertos (foco de gestión)
    porsub = defaultdict(list)
    for i in p1:
        porsub[i["sub"]].append(i)
    top = sorted(((s, resumir(v)) for s, v in porsub.items()),
                 key=lambda x: (-x[1]["atrasados"], -x[1]["abiertos"]))[:12]
    punch["topSubsistemas"] = [{"sub": s, **r} for s, r in top if r["abiertos"] > 0]

    # ---------- control cruzado con el reporte gerencial ----------
    ids_rep = {s["id"] for s in subs}
    ids_fuera = {s["id"] for s in fuera}
    ids_punch = {i["sub"] for i in items}
    # Los componentes están en el reporte aunque no en el universo: no son
    # huérfanos. Sí se cuenta aparte su punch, porque si algún día tienen ítems
    # abiertos, dejarlos fuera del universo escondería trabajo pendiente.
    huerfanos = sorted(ids_punch - ids_rep - ids_fuera)
    punch_fuera = sum(1 for i in items if i["sub"] in ids_fuera)
    control = {
        "punchTotal": len(items),
        "subsEnReporte": len(ids_rep),
        "subsConPunch": len(ids_punch & ids_rep),
        "subsSinPunch": len(ids_rep - ids_punch),
        "punchSinSubsistemaEnReporte": huerfanos[:20],
        "nHuerfanos": len(huerfanos),
        "itemsHuerfanos": sum(1 for i in items if i["sub"] in set(huerfanos)),
        "filasReporte": len(todos),
        "fueraDelUniverso": len(fuera),
        "punchEnFueraDelUniverso": punch_fuera,
        "contraste": [],
    }
    if punch_fuera:
        avisos.append(
            f"{punch_fuera} ítem(s) de punch están asignados a componentes, que quedan fuera "
            f"del universo — revisar si deben contarse en el subsistema padre")
    # Contraste del punch: reporte gerencial (tally por subsistema) vs punch list
    # (fuente ítem a ítem). Manda el punch list; la diferencia se informa.
    rep = declarado.get("_punch", {})
    if rep:
        difs = []
        for cat in ("P0", "P1", "P2", "P3"):
            r = rep.get(cat, {"abiertos": 0, "cerrados": 0})
            p = punch["porCategoria"].get(cat, {"abiertos": 0, "cerrados": 0})
            if r["abiertos"] != p["abiertos"] or r["cerrados"] != p["cerrados"]:
                difs.append({
                    "cat": cat,
                    "reporte": f"{r['cerrados']}C / {r['abiertos']}A",
                    "punch": f"{p['cerrados']}C / {p['abiertos']}A",
                })
        control["difPunch"] = difs
        if difs:
            avisos.append(
                "El reporte gerencial y el punch list no cuadran en "
                + ", ".join(d["cat"] for d in difs)
                + " — el panel usa el punch list, que es la fuente ítem a ítem")

    # Contraste de caminatas contra los totales que declara el propio archivo.
    for tipo, d in declarado.items():
        if tipo.startswith("_"):
            continue
        vig = camPorTipo.get(tipo, {}).get("vigente")
        if not vig:
            continue
        calc = camPorTipo[tipo][vig]
        ok = (calc["aplica"] == d["total"] and calc["realizada"] == d["ejecutadas"])
        control["contraste"].append({
            "tipo": tipo, "caminata": vig,
            "declarado": f"{d['ejecutadas']}/{d['total']}",
            "calculado": f"{calc['realizada']}/{calc['aplica']}",
            "ok": ok,
        })
        if not ok:
            avisos.append(
                f"Caminata {vig} de {tipo}: el archivo declara {d['ejecutadas']}/{d['total']} "
                f"pero el detalle da {calc['realizada']}/{calc['aplica']} — revisar")

    # ---------- semáforo por zona ----------
    semaforo = {}
    for z in zonas:
        ids = {s["id"] for s in subs if s["zona"] == z}
        p1z = [i for i in p1 if i["sub"] in ids]
        r = resumir(p1z)
        pct = 100 * r["cerrados"] / r["total"] if r["total"] else 100
        if r["atrasados"] > 40 or pct < 50:
            semaforo[z] = ["crit", "Crítico"]
        elif r["atrasados"] > 10 or pct < 80:
            semaforo[z] = ["warn", "Atención"]
        else:
            semaforo[z] = ["good", "Al día"]
        semaforo[z].append(r)

    return {
        "meta": {
            **PROYECTO,
            "corte": corte.strftime("%Y-%m-%d") if corte else None,
            "corteTexto": corte.strftime("%d-%m-%Y") if corte else "sin fecha",
            "hoy": hoy.strftime("%d-%m-%Y"),
            "generado": datetime.now().strftime("%Y-%m-%d %H:%M"),
        },
        "subsistemas": {
            "total": len(subs),
            "porTipo": dict(tipos),
            "porZona": dict(zonas),
            "zonaNombres": {z: ZONA_NOMBRES.get(z, z) for z in zonas},
            "filasReporte": len(todos),
        },
        "componentes": componentes,
        "caminatas": caminatas,
        "caminatasPorTipo": camPorTipo,
        "caminatasPorZona": camPorZona,
        "caminatasPendientes": pendientes_cam,
        "carpetas": carpetas,
        "punch": punch,
        "semaforo": semaforo,
        "control": control,
        "disciplinas": DISCIPLINAS,
        "discCorta": DISC_CORTA,
    }


MARCA_INI = "<!-- === DESALADORA:INICIO ==="
MARCA_FIN = "<!-- === DESALADORA:FIN === -->"


def inyectar_en_html(ruta_html, datos):
    """Reemplaza el bloque de datos de Desaladora dentro de index.html.

    Se inyecta en el propio HTML (y no en un .js aparte) para que el panel siga
    siendo UN archivo portable: se puede enviar por correo tal cual.
    """
    if not ruta_html.exists():
        avisos.append(f"No se encontró {ruta_html.name}; solo se escribió el JSON.")
        return False
    html = ruta_html.read_text(encoding="utf-8")
    i, f = html.find(MARCA_INI), html.find(MARCA_FIN)
    if i == -1 or f == -1 or f < i:
        avisos.append(f"No se encontraron las marcas DESALADORA en {ruta_html.name}; "
                      f"el panel quedó con los datos anteriores.")
        return False
    bloque = (
        MARCA_INI + " (bloque generado por desaladora.py — no editar a mano) -->\n"
        "<script>\nconst DES = "
        + json.dumps(datos, ensure_ascii=False, indent=1)
        + ";\n</script>\n"
    )
    ruta_html.write_text(html[:i] + bloque + html[f:], encoding="utf-8")
    return True


# =============================================================================
def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--reporte", required=True, help="REPORTE_GERENCIAL_*.xlsx")
    ap.add_argument("--punch", required=True, help="Listado_Puntos_Punch_Consolidado_*.xlsx")
    ap.add_argument("--hoy", help="Fecha de referencia AAAA-MM-DD (por defecto, hoy)")
    args = ap.parse_args()

    hoy = datetime.strptime(args.hoy, "%Y-%m-%d") if args.hoy else \
        datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    corte, subs, declarado = leer_reporte(args.reporte)
    items = leer_punch(args.punch, hoy)
    datos = construir(corte, subs, items, hoy, declarado)

    destino = AQUI / "datos_desaladora.json"
    destino.write_text(json.dumps(datos, ensure_ascii=False, indent=1), encoding="utf-8")
    inyectado = inyectar_en_html(AQUI / "index.html", datos)

    # ---------------- resumen en pantalla ----------------
    pc = lambda a, b: f"{pct1(a, b):.1f}%".replace(".", ",") if b else "—"
    m, s, c, cp, p = (datos["meta"], datos["subsistemas"], datos["caminatas"],
                      datos["carpetas"], datos["punch"])

    print("=" * 68)
    print(f"  {m['nombre'].upper()} — {m['descripcion']}")
    print(f"  {m['cliente']}   ·   contrato {m['contrato']}")
    print(f"  Corte del reporte: {m['corteTexto']}   ·   atrasos calculados al {m['hoy']}")
    print("=" * 68)

    print(f"\nSUBSISTEMAS: {s['total']}")
    print("  por tipo:", " · ".join(f"{k} {v}" for k, v in s["porTipo"].items()))
    print("  por zona:", " · ".join(f"{k} {v}" for k, v in s["porZona"].items()))
    comp = datos.get("componentes", {})
    if comp.get("total"):
        print(f"  fuera del universo: {comp['total']} componente(s), "
              f"{comp['entregados']} entregado(s) — el reporte trae {s['filasReporte']} filas")
        for cx in comp["detalle"]:
            print(f"    {cx['id']:24s} {cx['zona']:11s} carpeta {cx['carpeta']:14s} "
                  f"caminatas {','.join(map(str, cx['caminatas'])) or '—':5s} "
                  f"tarjeta verde {'SI' if cx['tarjetaVerde'] else 'no':3s} "
                  f"· cuelga de {cx['padre'] or '(sin padre en el reporte)'}")

    print("\nCAMINATAS")
    for n in ("1", "2", "3"):
        d = c[n]
        if d["aplica"]:
            na = f"  ({d['noAplica']} no aplican)" if d["noAplica"] else ""
            print(f"  Caminata {n}: {d['realizada']}/{d['aplica']} realizadas "
                  f"({pc(d['realizada'], d['aplica'])}){na}")
    print("  desglose por tipo de subsistema:")
    for t, v in datos["caminatasPorTipo"].items():
        vig = v["vigente"]
        if vig:
            d = v[vig]
            print(f"    {t:12s} (n={v['subs']:3d}) → caminata vigente {vig}: "
                  f"{d['realizada']}/{d['aplica']} ({pc(d['realizada'], d['aplica'])})")

    print(f"\nCARPETAS (certificado de entrega) — {cp['total']} subsistemas")
    for k in CARPETA_ORDEN:
        if cp["estados"][k]:
            print(f"  {k:16s} {cp['estados'][k]:3d}")
    print(f"  → entregadas al cliente: {cp['entregadas']}/{cp['total']} "
          f"({pc(cp['entregadas'], cp['total'])}) · tarjeta verde: {cp['tarjetaVerde']}")

    g = p["global"]
    print(f"\nPUNCH LIST — {g['total']} ítems")
    print(f"  Cerrados  {g['cerrados']:4d}  ({pc(g['cerrados'], g['total'])})")
    print(f"  Abiertos  {g['abiertos']:4d}  →  atrasados {g['atrasados']} · "
          f"en plazo {g['enPlazo']} · sin fecha requerida {g['sinFecha']}")
    print("\n  por categoría:")
    print(f"    {'':6s} {'total':>6s} {'cerr.':>6s} {'abie.':>6s} {'atras.':>7s}  {'% cierre':>9s}")
    for k in p["categorias"]:
        r = p["porCategoria"][k]
        if r["total"]:
            print(f"    {k:6s} {r['total']:6d} {r['cerrados']:6d} {r['abiertos']:6d} "
                  f"{r['atrasados']:7d}  {pc(r['cerrados'], r['total']):>9s}")
    print("\n  P1 por disciplina (abiertos / atrasados):")
    for d, r in sorted(p["p1PorDisciplina"].items(), key=lambda x: -x[1]["atrasados"]):
        print(f"    {d:26s} {r['total']:4d} tot · {r['abiertos']:3d} abiertos · "
              f"{r['atrasados']:3d} atrasados")
    a = p["antiguedad"]
    if a["max"]:
        print(f"\n  Antigüedad de los atrasados: mediana {a['mediana']} días · "
              f"máximo {a['max']} días")
        print("   ", " · ".join(f"{k}: {v}" for k, v in a["tramos"].items() if v))

    ctl = datos["control"]
    print(f"\nCONTROL DE CRUCE")
    print(f"  Subsistemas en el reporte: {ctl['subsEnReporte']} · "
          f"con punch asociado: {ctl['subsConPunch']} · sin punch: {ctl['subsSinPunch']}")
    for c in ctl["contraste"]:
        marca = "OK " if c["ok"] else "≠  "
        print(f"  {marca} Caminata {c['caminata']} de {c['tipo']}: "
              f"archivo declara {c['declarado']} · calculado {c['calculado']}")
    if ctl["nHuerfanos"]:
        print(f"  ⚠ {ctl['nHuerfanos']} subsistemas del punch no existen en el reporte "
              f"gerencial ({ctl['itemsHuerfanos']} ítems): "
              f"{', '.join(ctl['punchSinSubsistemaEnReporte'][:10])}")
    if ctl.get("difPunch"):
        print("  ⚠ El conteo de punch difiere entre los dos archivos "
              "(el panel usa el punch list):")
        for d in ctl["difPunch"]:
            print(f"      {d['cat']}: reporte gerencial {d['reporte']} · "
                  f"punch list {d['punch']}")

    if avisos:
        print("\n⚠  AVISOS — revisar antes de publicar:")
        for x in sorted(set(avisos))[:20]:
            print(f"   · {x}")

    print(f"\nEscritos: {destino.name}" + (", index.html (pestaña Desaladora)" if inyectado else ""))


if __name__ == "__main__":
    main()
