#!/usr/bin/env python3
"""
Procesa el proyecto TALABRE (Codelco) y escribe datos_talabre.json, que alimenta
la pestaña Talabre del panel y la PPT.

Uso (cada semana, con los dos archivos del corte):

    python3 talabre.py \
        --status /ruta/STATUS_SUBSISTEMAS_TALABRE.xlsx \
        --dt     /ruta/Detalle_de_TerminacionesBesalco.xlsx

Opcional:
    --hoy 2026-08-03     fecha de referencia para verificar los atrasos
                         (por defecto, el día en que se corre el script)

------------------------------------------------------------------------------
DE DÓNDE SALE CADA DATO
------------------------------------------------------------------------------
Las columnas se resuelven POR NOMBRE DE ENCABEZADO, no por posición, y la fila
del encabezado se detecta sola. Los dos archivos ya cambiaron de layout una vez
—el de subsistemas insertó tres columnas de protocolos y el de DT pasó su
encabezado a la fila 9—, así que fijar índices los volvería a romper.

STATUS · hoja «STATUS» — se buscan estos encabezados:
    ÁREA · SUBSISTEMAS · RESPONSABLE · CAMINATA 1 · CAMINATA 2 · ALCANCE ·
    % PEC · DT P1 · DT P2 · PENDIENTES CONSTRUCTIVO
    La columna ÁREA viene con celdas combinadas: el valor se arrastra hacia
    abajo hasta que aparece otro.
STATUS · hoja «RESUMEN» — totales que declara el propio archivo, contra los que
    se contrasta el cálculo. Es OPCIONAL: si no viene, se pierde ese contraste
    y el script avisa.

DT · hoja «DT» — se buscan estos encabezados:
    N° SUBSISTEMA · NOMBRE SISTEMA · ÁREA · PRIORIDADES · FECHA DE EMISION ·
    FECHA DE COMPROMISO DE CIERRE · FECHA DE CIERRE · ID EXTERNO · CAMINATA ·
    DISC.   (y, si el archivo las trae, STATUS y DIAS ATRASO)

------------------------------------------------------------------------------
REGLAS DE CLASIFICACIÓN
------------------------------------------------------------------------------
ESTADO del DT — se calcula con la MISMA regla que el resto de los proyectos:
        Cerrado  = tiene fecha de cierre
        Atrasado = sin fecha de cierre y con fecha de compromiso ya vencida
        Abierto  = el resto
    Los abiertos sin fecha de compromiso no se dan por atrasados: se informan
    aparte. Si el archivo trae su propia columna STATUS (el registro anterior la
    traía), manda esa y el script avisa de los registros donde no coincidan.

ÁREAS — el panel trabaja con las 9 áreas del proyecto. Los dos archivos abren
    los pozos uno a uno (PBO-15, PBBR-02, TB-01…): se agrupan en POZOS, igual
    que hacía la hoja STATUS antes de abrirlos. «ADUCCIÓN» se rotula
    «IMPULSIÓN ADUCCIÓN». LÍNEAS IMPULSIÓN y PQS son áreas propias aunque
    cuelguen del sistema «Pozos Barrera Hidráulica»: agruparlas en POZOS —como
    hacía la versión anterior, que leía el sistema y no el área— escondía 296
    detalles bajo un área que no es la suya.

CAMINATAS — se identifican por NÚMERO. En la hoja STATUS:
    «SI» = realizada · «NO» o vacío = sin programar
    una fecha AÚN POR LLEGAR = programada en plazo
    una fecha YA PASADA = vencida (se agendó y no se hizo: sí es incumplimiento)
    OJO: la hoja RESUMEN publica «C1/C2 Prog.», que suma las realizadas MÁS las
    programadas de la semana. Este panel las informa por separado, porque una
    caminata agendada no es una caminata hecha. El contraste queda documentado
    en el control de calidad del dato.

CARPETAS — Talabre no maneja estados de aprobación como los otros proyectos:
    lleva un % de avance por subsistema (% PEC). Se reporta como avance, con la
    distribución por tramos, y NO se traduce a «entregadas» para no inventar un
    concepto que el proyecto no usa. El propio archivo declara además cuántas
    están en revisión del cliente.

DISCIPLINAS — se unifican los nombres del proyecto con los del resto:
    CANERIAS -> Piping · CIVIL y MOVIMIENTO DE TIERRAS -> Obras Civiles
    ELECTRICOS -> Eléctrica · INSTRUMENTACION -> Instrumentación y Control
"""

import argparse
import json
import re
import sys
import unicodedata
import warnings
from collections import Counter, defaultdict
from datetime import datetime
from math import floor
from pathlib import Path

warnings.filterwarnings("ignore")

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Falta openpyxl.  Instálalo con:  pip install openpyxl")

AQUI = Path(__file__).resolve().parent

PROYECTO = {
    "id": "TALABRE",
    "nombre": "Talabre",
    "descripcion": "Captación de agua, pozos y estaciones de bombeo",
    # Código y cliente tal como los escribe el módulo de Protocolos.
    "codigo": "P2407",
    "cliente": "CODELCO",
    "contrato": "4600029355",
}

DISCIPLINAS = ["Obras Civiles", "Piping", "Eléctrica",
               "Instrumentación y Control", "Estructura"]
DISC_CORTA = {"Instrumentación y Control": "Instr. y Control"}

DISC_MAP = {
    "civil": "Obras Civiles",
    "movimiento de tierras": "Obras Civiles",
    "oo.cc": "Obras Civiles",
    "canerias": "Piping",
    "piping": "Piping",
    "electricos": "Eléctrica",
    "electrica": "Eléctrica",
    "instrumentacion": "Instrumentación y Control",
    "instrumentacion y control": "Instrumentación y Control",
    "estructura": "Estructura",
}

# Estados que trae la propia columna STATUS del cuadro de DT.
CERRADO, ABIERTO, ATRASADO = "Cerrado", "Abierto", "Atrasado"

# Tramos del % PEC de la carpeta.
TRAMOS_PEC = [("95% o más", 0.95), ("80% a 95%", 0.80), ("Bajo 80%", 0.0)]

avisos = []


def norm(v):
    """Minúsculas, sin tildes, sin espacios sobrantes."""
    if v is None:
        return ""
    s = str(v).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    return " ".join(s.split())


def disciplina(v):
    d = DISC_MAP.get(norm(v))
    if d:
        return d
    if v not in (None, ""):
        avisos.append(f"Disciplina no reconocida: «{v}» — se agrupa en «Otras»")
    return "Otras"


def prioridad(v):
    """Acepta «P1», «P1A»… y también el dígito suelto («1», 1) del registro nuevo."""
    s = norm(v).upper().replace(" ", "")
    m = re.match(r"^P?(\d)", s)
    if m:
        return "P" + m.group(1)
    if s in ("", "SINASIGNAR"):
        return "Sin asignar"
    avisos.append(f"Prioridad no reconocida: «{v}» — se agrupa en «Sin asignar»")
    return "Sin asignar"


def pct1(a, b):
    """Porcentaje con 1 decimal, redondeando .5 hacia arriba (igual que el panel)."""
    if not b:
        return 0.0
    return floor(1000.0 * a / b + 0.5) / 10.0


def as_fecha(v):
    if isinstance(v, datetime):
        return v
    if isinstance(v, str) and v.strip():
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(v.strip(), fmt)
            except ValueError:
                pass
    return None


# Las dos fuentes abren los pozos uno a uno (PBO-15, PBBR-02, PBN-01, TB-01…).
# El panel los agrupa en POZOS, que es el área con la que siempre se reportó.
POZOS_RE = re.compile(r"^(pbo|pbbr|pbn|tb)[\s-]")


def area_canon(v):
    """Área tal como la reporta el panel: pozos agrupados y rótulos unificados."""
    if v in (None, ""):
        return "Sin área"
    a = str(v).strip()
    na = norm(a)
    if na == "aduccion":
        return "IMPULSIÓN ADUCCIÓN"
    if POZOS_RE.match(na):
        return "POZOS"
    return a


def encabezado(ws, requeridas, max_filas=15):
    """Busca la fila que contiene todas las columnas requeridas y devuelve
    (n_fila, {nombre_norm: índice}). Los dos archivos ya movieron su encabezado
    de fila una vez; fijarlo a mano los volvería a romper."""
    for nf, fila in enumerate(ws.iter_rows(min_row=1, max_row=max_filas,
                                           values_only=True), 1):
        cols = {}
        for i, v in enumerate(fila):
            if v not in (None, ""):
                cols[norm(v).rstrip(".")] = i
        if all(r in cols for r in requeridas):
            return nf, cols
    sys.exit(f"No se encontró el encabezado con {requeridas} en la hoja «{ws.title}»")


# =============================================================================
# 1) STATUS — subsistemas, caminatas y avance de carpeta
# =============================================================================
def leer_status(ruta, hoy):
    wb = load_workbook(ruta, data_only=True)
    ws = wb["STATUS"]
    nf, c = encabezado(ws, ["area", "subsistemas", "caminata 1", "caminata 2", "% pec"])
    col = lambda *nombres: next((c[n] for n in nombres if n in c), None)
    iSub, iArea = c["subsistemas"], c["area"]
    iC1, iC2, iPec = c["caminata 1"], c["caminata 2"], c["% pec"]
    iResp, iAlc = col("responsable"), col("alcance")
    iP1, iP2 = col("dt p1"), col("dt p2")
    iObs = col("pendientes constructivo", "observaciones")

    def estado_cam(v):
        """SI = realizada · fecha por llegar = programada en plazo ·
        fecha ya pasada = vencida (se agendó y no se hizo) · NO o vacío = sin programar."""
        f = v if isinstance(v, datetime) else as_fecha(v)
        if f:
            return ("Programada", f) if f >= hoy else ("Vencida", f)
        s = norm(v)
        if s in ("si", "sí"):
            return "Realizada", None
        if s in ("no", ""):
            return "Pendiente", None
        avisos.append(f"Estado de caminata no reconocido: «{v}» — se cuenta como pendiente")
        return "Pendiente", None

    subs, corte, area = [], None, "Sin área"
    for f in ws.iter_rows(min_row=nf + 1, values_only=True):
        if f[iSub] in (None, ""):
            continue
        if f[iArea] not in (None, ""):        # la columna viene con celdas combinadas
            area = area_canon(f[iArea])
        e1, d1 = estado_cam(f[iC1])
        e2, d2 = estado_cam(f[iC2])
        for d in (d1, d2):
            if d and (corte is None or d > corte):
                corte = d
        subs.append({
            "id": str(f[iSub]).strip(),
            "area": area,
            "resp": str(f[iResp]).strip() if iResp is not None and f[iResp] else "",
            "alcance": str(f[iAlc]).strip() if iAlc is not None and f[iAlc] else "",
            "cam": {1: e1, 2: e2},
            "fechaCam": {1: d1.strftime("%d-%m-%Y") if d1 else None,
                         2: d2.strftime("%d-%m-%Y") if d2 else None},
            "pec": f[iPec] if isinstance(f[iPec], (int, float)) else None,
            "dtP1": int(f[iP1]) if iP1 is not None and isinstance(f[iP1], (int, float)) else 0,
            "dtP2": int(f[iP2]) if iP2 is not None and isinstance(f[iP2], (int, float)) else 0,
            "obs": str(f[iObs]).strip() if iObs is not None and f[iObs] else "",
        })

    # Totales que declara el propio archivo, para contrastarlos. La hoja RESUMEN
    # dejó de venir en el archivo nuevo: si no está, se pierde ese contraste.
    declarado = {"areas": {}, "total": None}
    if "RESUMEN" in wb.sheetnames:
        for r in wb["RESUMEN"].iter_rows(min_row=3, max_row=12, values_only=True):
            vals = [r[i] for i in range(2, 8)]
            if not any(isinstance(v, (int, float)) for v in vals):
                continue
            fila = {"subs": r[2], "c1": r[3], "c2": r[4],
                    "sobre80": r[5], "sobre95": r[6], "enRevision": r[7]}
            if r[1]:
                declarado["areas"][str(r[1]).strip()] = fila
            else:
                declarado["total"] = fila
    else:
        avisos.append("El archivo de STATUS ya no trae la hoja «RESUMEN»: se pierde "
                      "el contraste contra los totales que declaraba el proyecto")

    # La hoja «DT» del mismo archivo es el extracto de pendientes: sirve para
    # contrastar cuántos abiertos declara el proyecto.
    declarado["dtAbiertos"] = None
    if "DT" in wb.sheetnames:
        declarado["dtAbiertos"] = sum(
            1 for r in wb["DT"].iter_rows(min_row=2, values_only=True)
            if r[0] not in (None, ""))
    return corte, subs, declarado


# =============================================================================
# 2) Registro de DT — detalles de terminación
# =============================================================================
def leer_dt(ruta, hoy):
    wb = load_workbook(ruta, data_only=True, read_only=True)
    ws = wb["DT"]
    nf, c = encabezado(ws, ["n° subsistema", "prioridades", "caminata", "disc",
                            "fecha de compromiso de cierre", "fecha de cierre"])
    col = lambda *nombres: next((c[n] for n in nombres if n in c), None)
    iSub, iPrio, iCam, iDisc = c["n° subsistema"], c["prioridades"], c["caminata"], c["disc"]
    iComp, iCierre = c["fecha de compromiso de cierre"], c["fecha de cierre"]
    iArea, iId = col("área", "area"), col("id externo")
    iSt, iDias = col("status"), col("dias atraso")
    if iSt is None:
        avisos.append("El registro de DT no trae columna STATUS: el estado se calcula "
                      "con la regla homologada (cerrado = con fecha de cierre; "
                      "atrasado = abierto con compromiso vencido)")

    items, incoherentes = [], []
    for f in ws.iter_rows(min_row=nf + 1, values_only=True):
        if f[iSub] in (None, "") and (iId is None or f[iId] in (None, "")):
            continue
        comp, cierre = as_fecha(f[iComp]), as_fecha(f[iCierre])

        if iSt is not None:
            # El archivo trae su STATUS: manda él, pero se verifica que respete
            # la misma regla que usamos en los otros proyectos.
            st = str(f[iSt]).strip() if f[iSt] else ""
            if st not in (CERRADO, ABIERTO, ATRASADO):
                avisos.append(f"STATUS de DT no reconocido: «{f[iSt]}» — se cuenta como abierto")
                st = ABIERTO
            rid = str(f[iId]) if iId is not None else str(f[iSub])
            if st == ATRASADO and not (cierre is None and comp and comp < hoy):
                incoherentes.append(("marcado Atrasado sin compromiso vencido", rid))
            if st == ABIERTO and cierre is None and comp and comp < hoy:
                incoherentes.append(("abierto con compromiso vencido pero no marcado", rid))
            if st == CERRADO and cierre is None:
                incoherentes.append(("marcado Cerrado sin fecha de cierre", rid))
        else:
            st = CERRADO if cierre else (ATRASADO if comp and comp < hoy else ABIERTO)

        if iDias is not None and isinstance(f[iDias], (int, float)) and st == ATRASADO:
            dias = int(f[iDias])
        else:
            dias = (hoy - comp).days if st == ATRASADO and comp else None

        cam = re.sub(r"\D", "", str(f[iCam] or ""))
        items.append({
            "id": str(f[iId]).strip() if iId is not None and f[iId] else str(f[iSub]),
            "sub": str(f[iSub]).strip() if f[iSub] else "",
            "area": area_canon(f[iArea]) if iArea is not None else "Sin área",
            "disc": disciplina(f[iDisc]),
            "prio": prioridad(f[iPrio]),
            "estado": st,
            "cam": int(cam) if cam and int(cam) in (1, 2, 3) else 0,
            "diasAtraso": dias,
            "sinCompromiso": st in (ABIERTO, ATRASADO) and comp is None,
        })

    if incoherentes:
        c = Counter(m for m, _ in incoherentes)
        for m, n in c.items():
            avisos.append(f"{n} DT con estado inconsistente: {m}")
    return items, incoherentes, iSt is not None


# =============================================================================
# 3) Agregaciones
# =============================================================================
def resumir(items):
    ab = [i for i in items if i["estado"] in (ABIERTO, ATRASADO)]
    atr = [i for i in ab if i["estado"] == ATRASADO]
    return {
        "total": len(items),
        "cerrados": sum(1 for i in items if i["estado"] == CERRADO),
        "abiertos": len(ab),
        "atrasados": len(atr),
        "enPlazo": len(ab) - len(atr),
        "sinCompromiso": sum(1 for i in ab if i["sinCompromiso"]),
    }


def construir(corte, subs, declarado, items, incoherentes, hoy, estado_propio):
    areas = sorted({s["area"] for s in subs})

    # ---------- caminatas ----------
    # «Programada» es solo la agendada a una fecha AÚN POR LLEGAR. La que se
    # agendó y la fecha ya pasó sin marcarse realizada queda «Vencida»: esa sí
    # es exigible, porque descontarla escondería un incumplimiento real.
    caminatas = {}
    for n in (1, 2):
        caminatas[str(n)] = {
            "total": len(subs),
            "realizada": sum(1 for s in subs if s["cam"][n] == "Realizada"),
            "programada": sum(1 for s in subs if s["cam"][n] == "Programada"),
            "vencida": sum(1 for s in subs if s["cam"][n] == "Vencida"),
            "pendiente": sum(1 for s in subs if s["cam"][n] == "Pendiente"),
        }
        c = caminatas[str(n)]
        c["pct"] = pct1(c["realizada"], c["total"])
        # Avance sobre lo EXIGIBLE: se descuentan las agendadas a fecha futura,
        # que no son un incumplimiento al corte. Se guardan las dos cifras —la
        # del universo y la de lo exigible— porque miden cosas distintas y el
        # panel las muestra juntas: descontar no es esconder.
        c["exigible"] = c["total"] - c["programada"]
        c["pctExigible"] = pct1(c["realizada"], c["exigible"]) if c["exigible"] else 100.0

    camArea = {}
    for a in areas:
        sa = [s for s in subs if s["area"] == a]
        camArea[a] = {"subs": len(sa)}
        for n in (1, 2):
            camArea[a][str(n)] = {
                "total": len(sa),
                "realizada": sum(1 for s in sa if s["cam"][n] == "Realizada"),
                "programada": sum(1 for s in sa if s["cam"][n] == "Programada"),
                "vencida": sum(1 for s in sa if s["cam"][n] == "Vencida"),
                "pendiente": sum(1 for s in sa if s["cam"][n] == "Pendiente"),
                "exigible": len(sa) - sum(1 for s in sa if s["cam"][n] == "Programada"),
            }

    pendientes = sorted(
        [{"id": s["id"], "area": s["area"], "alcance": s["alcance"],
          "cam": n, "estado": s["cam"][n], "fecha": s["fechaCam"][n], "pec": s["pec"]}
         for s in subs for n in (2, 1) if s["cam"][n] != "Realizada"],
        key=lambda x: (x["cam"], x["estado"], x["area"], x["id"]))

    # ---------- carpetas (% PEC) ----------
    pecs = [s["pec"] for s in subs if isinstance(s["pec"], (int, float))]
    tramos = {}
    for i, (nom, piso) in enumerate(TRAMOS_PEC):
        techo = TRAMOS_PEC[i - 1][1] if i else 1.01
        tramos[nom] = sum(1 for p in pecs if piso <= p < techo)
    carpetas = {
        "tipo": "avance",
        "total": len(subs),
        "conDato": len(pecs),
        "promedio": pct1(sum(pecs), len(pecs)) if pecs else 0,
        "sobre95": sum(1 for p in pecs if p >= 0.95),
        "sobre80": sum(1 for p in pecs if p >= 0.80),
        "bajo80": sum(1 for p in pecs if p < 0.80),
        "tramos": tramos,
        "enRevisionCliente": (declarado.get("total") or {}).get("enRevision"),
    }
    carpetas["porArea"] = {}
    for a in areas:
        pa = [s["pec"] for s in subs if s["area"] == a and isinstance(s["pec"], (int, float))]
        carpetas["porArea"][a] = {
            "subs": sum(1 for s in subs if s["area"] == a),
            "promedio": pct1(sum(pa), len(pa)) if pa else 0,
            "sobre95": sum(1 for p in pa if p >= 0.95),
            "sobre80": sum(1 for p in pa if p >= 0.80),
            "bajo80": sum(1 for p in pa if p < 0.80),
        }
    carpetas["rezagados"] = sorted(
        [{"id": s["id"], "area": s["area"], "pec": round(100 * s["pec"], 1), "obs": s["obs"]}
         for s in subs if isinstance(s["pec"], (int, float)) and s["pec"] < 0.80],
        key=lambda x: x["pec"])

    # ---------- DT ----------
    prios = ["P1", "P2", "P3", "P4", "Sin asignar"]
    p1 = [i for i in items if i["prio"] == "P1"]
    areasDT = sorted({i["area"] for i in items})

    dt = {
        "global": resumir(items),
        "porPrioridad": {p: resumir([i for i in items if i["prio"] == p]) for p in prios},
        "porDisciplina": {d: resumir([i for i in items if i["disc"] == d])
                          for d in DISCIPLINAS if any(i["disc"] == d for i in items)},
        "p1PorDisciplina": {d: resumir([i for i in p1 if i["disc"] == d])
                            for d in DISCIPLINAS if any(i["disc"] == d for i in p1)},
        "p1PorArea": {a: resumir([i for i in p1 if i["area"] == a]) for a in areasDT},
        "porArea": {a: resumir([i for i in items if i["area"] == a]) for a in areasDT},
        "porCaminata": {},
        "prioridades": [p for p in prios if any(i["prio"] == p for i in items)],
    }
    for n in (1, 2, 3):
        sub = [i for i in items if i["cam"] == n]
        if sub:
            dt["porCaminata"][str(n)] = resumir(sub)

    # heatmap: abiertos (P1) por disciplina × caminata
    dt["heatmap"] = {
        "disc": [d for d in DISCIPLINAS if any(i["disc"] == d for i in p1)],
        "cam": [str(n) for n in (1, 2, 3) if any(i["cam"] == n for i in p1)],
    }
    dt["heatmap"]["vals"] = [
        [sum(1 for i in p1 if i["disc"] == d and i["cam"] == int(c)
             and i["estado"] in (ABIERTO, ATRASADO))
         for c in dt["heatmap"]["cam"]]
        for d in dt["heatmap"]["disc"]]

    # matriz disciplina × área de DT abiertos (alimenta el heatmap del panel)
    dt["matriz"] = {
        d: {a: sum(1 for i in items if i["disc"] == d and i["area"] == a
                   and i["estado"] in (ABIERTO, ATRASADO))
            for a in areasDT}
        for d in dt["porDisciplina"]}

    dias = sorted(i["diasAtraso"] for i in items if i["diasAtraso"])
    tr = Counter(">180 días" if d > 180 else "91-180 días" if d > 90
                 else "31-90 días" if d > 30 else "1-30 días" for d in dias)
    dt["antiguedad"] = {
        "tramos": {k: tr.get(k, 0) for k in
                   ["1-30 días", "31-90 días", "91-180 días", ">180 días"]},
        "max": dias[-1] if dias else 0,
        "mediana": dias[len(dias) // 2] if dias else 0,
    }

    porsub = defaultdict(list)
    for i in items:
        if i["estado"] in (ABIERTO, ATRASADO):
            porsub[i["sub"]].append(i)
    top = sorted(((s, resumir(v)) for s, v in porsub.items()),
                 key=lambda x: (-x[1]["atrasados"], -x[1]["abiertos"]))[:12]
    dt["topSubsistemas"] = [{"sub": s, **r} for s, r in top]

    # ---------- control de calidad del dato ----------
    # estadoPropio: si el registro trae su columna STATUS, la regla de atraso es
    # una VERIFICACIÓN; si no la trae, el estado se calcula con la regla y el
    # panel lo declara en vez de presentar un chequeo que no puede fallar.
    control = {"contraste": [], "incoherentes": len(incoherentes),
               "estadoPropio": estado_propio}
    tot = declarado.get("total") or {}
    if tot.get("subs") is not None:
        control["contraste"].append({
            "que": "Subsistemas", "declarado": str(tot["subs"]),
            "calculado": str(len(subs)), "ok": tot["subs"] == len(subs)})
    for n, key in ((1, "c1"), (2, "c2")):
        if tot.get(key) is not None:
            real = caminatas[str(n)]["realizada"]
            prog = caminatas[str(n)]["programada"]
            control["contraste"].append({
                "que": f"Caminata {n} — el RESUMEN publica «programadas»",
                "declarado": f"{tot[key]} programadas",
                "calculado": f"{real} realizadas + {prog} programadas",
                "ok": tot[key] <= real + prog,
                "nota": "el RESUMEN suma las agendadas de la semana; el panel las separa"})
    if tot.get("sobre95") is not None:
        control["contraste"].append({
            "que": "Carpetas sobre 95%", "declarado": str(tot["sobre95"]),
            "calculado": str(carpetas["sobre95"]), "ok": tot["sobre95"] == carpetas["sobre95"]})
    if tot.get("sobre80") is not None:
        ok = tot["sobre80"] == carpetas["sobre80"]
        control["contraste"].append({
            "que": "Carpetas sobre 80%", "declarado": str(tot["sobre80"]),
            "calculado": str(carpetas["sobre80"]), "ok": ok})
        if not ok:
            avisos.append(f"Carpetas sobre 80%: el RESUMEN declara {tot['sobre80']} "
                          f"y el detalle da {carpetas['sobre80']} — revisar")

    # La hoja «DT» del archivo de STATUS es el extracto de pendientes del
    # proyecto: debe cuadrar con los abiertos del registro completo.
    if declarado.get("dtAbiertos") is not None:
        control["contraste"].append({
            "que": "DT abiertos (hoja DT del archivo de STATUS)",
            "declarado": str(declarado["dtAbiertos"]),
            "calculado": str(dt["global"]["abiertos"]),
            "ok": declarado["dtAbiertos"] == dt["global"]["abiertos"]})
        if declarado["dtAbiertos"] != dt["global"]["abiertos"]:
            avisos.append(f"El extracto de pendientes trae {declarado['dtAbiertos']} DT y el "
                          f"registro completo da {dt['global']['abiertos']} abiertos — revisar "
                          "si los archivos son del mismo corte")

    # DT abiertos: la hoja STATUS lleva su propio conteo por subsistema.
    # Si cuadra con el registro se muestra como contraste OK; si no, va en
    # difDT, que el panel pinta como advertencia con su explicación.
    sP1 = sum(s["dtP1"] for s in subs)
    sP2 = sum(s["dtP2"] for s in subs)
    rP1 = dt["porPrioridad"]["P1"]["abiertos"]
    rP2 = dt["porPrioridad"]["P2"]["abiertos"]
    control["difDT"] = []
    for nom, sv, rv in (("P1", sP1, rP1), ("P2", sP2, rP2)):
        if sv != rv:
            control["difDT"].append({"prio": nom, "status": sv, "registro": rv})
        else:
            control["contraste"].append({
                "que": f"DT {nom} abiertos: hoja STATUS vs. registro",
                "declarado": str(sv), "calculado": str(rv), "ok": True})
    if control["difDT"]:
        avisos.append("La hoja STATUS y el registro de DT no cuadran en "
                      + ", ".join(d["prio"] for d in control["difDT"])
                      + " — el panel usa el registro de DT, que es la fuente ítem a ítem")

    ids_status = {s["id"] for s in subs}
    control["subsEnStatus"] = len(ids_status)
    control["areasStatus"] = len(areas)
    control["areasDT"] = len(areasDT)
    control["dtSinCompromiso"] = dt["global"]["sinCompromiso"]

    # ---------- semáforo por área ----------
    # El avance de caminata se mide sobre lo EXIGIBLE, no sobre el universo: una
    # caminata con fecha agendada todavía por llegar no es un incumplimiento hoy.
    # Las agendadas se descuentan del denominador y se informan aparte, para que
    # descontarlas no equivalga a esconderlas.
    #   exigible = total − agendadas   (realizadas + las que ya debieron estar)
    # Regla pedida por Mauricio Rocha (31-07-2026) y aprobada por el usuario.
    # Un área SIN BASE —sin ningún detalle levantado y sin ninguna caminata
    # vigente realizada— no se puntúa: no hay nada medido todavía y pintarla del
    # mismo rojo que un frente con atrasos reales desinforma.
    semaforo = {}
    for a in areas:
        c2 = camArea[a]["2"]
        r = dt["p1PorArea"].get(a) or resumir([])
        cp = carpetas["porArea"][a]
        agendadas = c2.get("programada", 0)
        exigible = c2["total"] - agendadas
        avanceCam = pct1(c2["realizada"], exigible) if exigible else 100.0
        cierre = pct1(r["cerrados"], r["total"]) if r["total"] else 100.0
        if r["total"] == 0 and c2["realizada"] == 0:
            est = ["nd", "Sin base al corte"]
        elif avanceCam < 60 or cierre < 50 or r["atrasados"] > 20:
            est = ["crit", "Crítico"]
        elif avanceCam < 90 or cierre < 90 or r["atrasados"] > 5 or cp["bajo80"]:
            est = ["warn", "Atención"]
        else:
            est = ["good", "Al día"]
        semaforo[a] = est + [r, {"exigible": exigible, "agendadas": agendadas,
                                 "avanceExigible": avanceCam}]

    return {
        "meta": {
            **PROYECTO,
            # El corte de Talabre es la fecha de referencia con la que se
            # calculan los atrasos. La caminata agendada más lejana se guarda
            # aparte: es una fecha futura y no debe confundirse con el corte.
            "corte": hoy.strftime("%Y-%m-%d"),
            "corteTexto": hoy.strftime("%d-%m-%Y"),
            "hoy": hoy.strftime("%d-%m-%Y"),
            "ultimaAgenda": corte.strftime("%d-%m-%Y") if corte else None,
            "generado": datetime.now().strftime("%Y-%m-%d %H:%M"),
        },
        "subsistemas": {"total": len(subs), "porArea": {a: sum(1 for s in subs if s["area"] == a)
                                                        for a in areas}},
        "areas": areas,
        "caminatas": caminatas,
        "caminatasPorArea": camArea,
        "caminatasPendientes": pendientes,
        "carpetas": carpetas,
        "dt": dt,
        "semaforo": semaforo,
        "control": control,
        "disciplinas": DISCIPLINAS,
        "discCorta": DISC_CORTA,
    }


MARCA_INI = "<!-- === TALABRE:INICIO ==="
MARCA_FIN = "<!-- === TALABRE:FIN === -->"


def inyectar_en_html(ruta_html, datos):
    """Reemplaza el bloque de datos de Talabre dentro de index.html."""
    if not ruta_html.exists():
        avisos.append(f"No se encontró {ruta_html.name}; solo se escribió el JSON.")
        return False
    html = ruta_html.read_text(encoding="utf-8")
    i, f = html.find(MARCA_INI), html.find(MARCA_FIN)
    if i == -1 or f == -1 or f < i:
        avisos.append(f"No se encontraron las marcas TALABRE en {ruta_html.name}; "
                      f"el panel quedó con los datos anteriores.")
        return False
    bloque = (
        MARCA_INI + " (bloque generado por talabre.py — no editar a mano) -->\n"
        "<script>\nconst TAL = "
        + json.dumps(datos, ensure_ascii=False, indent=1)
        + ";\n</script>\n"
    )
    ruta_html.write_text(html[:i] + bloque + html[f:], encoding="utf-8")
    return True


# =============================================================================
def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--status", required=True, help="STATUS_SUBSISTEMAS_TALABRE.xlsx")
    ap.add_argument("--dt", required=True, help="Detalle_de_TerminacionesBesalco.xlsx")
    ap.add_argument("--hoy", help="Fecha de referencia AAAA-MM-DD (por defecto, hoy)")
    args = ap.parse_args()

    hoy = datetime.strptime(args.hoy, "%Y-%m-%d") if args.hoy else \
        datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    corte, subs, declarado = leer_status(args.status, hoy)
    items, incoherentes, estado_propio = leer_dt(args.dt, hoy)
    datos = construir(corte, subs, declarado, items, incoherentes, hoy, estado_propio)

    destino = AQUI / "datos_talabre.json"
    destino.write_text(json.dumps(datos, ensure_ascii=False, indent=1), encoding="utf-8")
    inyectado = inyectar_en_html(AQUI / "index.html", datos)

    # ---------------- resumen en pantalla ----------------
    pc = lambda a, b: f"{pct1(a, b):.1f}%".replace(".", ",") if b else "—"
    m, S, C, CP, T = (datos["meta"], datos["subsistemas"], datos["caminatas"],
                      datos["carpetas"], datos["dt"])

    print("=" * 70)
    print(f"  {m['nombre'].upper()} — {m['descripcion']}")
    print(f"  {m['cliente']}   ·   contrato {m['contrato']}")
    print(f"  Corte: {m['corteTexto']}   ·   última caminata agendada: {m['ultimaAgenda'] or '—'}")
    print("=" * 70)

    print(f"\nSUBSISTEMAS: {S['total']} en {len(datos['areas'])} áreas")
    for a, n in sorted(S["porArea"].items(), key=lambda x: -x[1]):
        print(f"    {a:26s} {n:4d}")

    print("\nCAMINATAS")
    for n in ("1", "2"):
        c = C[n]
        print(f"  Caminata {n}: {c['realizada']}/{c['exigible']} exigibles ({pc(c['realizada'], c['exigible'])})"
              f" · {pc(c['realizada'], c['total'])} del universo · {c['programada']} programadas en plazo"
              f" · {c['vencida']} con agenda vencida · {c['pendiente']} sin programar")

    print(f"\nCARPETAS — avance PEC (Talabre no usa estados de aprobación)")
    print(f"  Promedio: {CP['promedio']}%".replace(".", ","))
    for k, v in CP["tramos"].items():
        print(f"    {k:12s} {v:4d}")
    if CP["enRevisionCliente"] is not None:
        print(f"  En revisión del cliente (según el archivo): {CP['enRevisionCliente']}")

    g = T["global"]
    print(f"\nDETALLES DE TERMINACIÓN — {g['total']} registros")
    print(f"  Cerrados  {g['cerrados']:5d}  ({pc(g['cerrados'], g['total'])})")
    print(f"  Abiertos  {g['abiertos']:5d}  →  atrasados {g['atrasados']} · "
          f"en plazo {g['enPlazo']} ({g['sinCompromiso']} sin fecha de compromiso)")
    print(f"\n  {'':12s} {'total':>6s} {'cerr.':>6s} {'abie.':>6s} {'atras.':>7s} {'% cierre':>9s}")
    for p in T["prioridades"]:
        r = T["porPrioridad"][p]
        print(f"  {p:12s} {r['total']:6d} {r['cerrados']:6d} {r['abiertos']:6d} "
              f"{r['atrasados']:7d} {pc(r['cerrados'], r['total']):>9s}")
    print("\n  P1 por disciplina:")
    for d, r in sorted(T["p1PorDisciplina"].items(), key=lambda x: -x[1]["atrasados"]):
        print(f"    {d:26s} {r['total']:4d} tot · {r['abiertos']:3d} abiertos · "
              f"{r['atrasados']:3d} atrasados")
    a = T["antiguedad"]
    if a["max"]:
        print(f"\n  Antigüedad de los atrasados: mediana {a['mediana']} días · máximo {a['max']} días")
        print("   ", " · ".join(f"{k}: {v}" for k, v in a["tramos"].items() if v))

    K = datos["control"]
    print("\nCONTROL DE CRUCE")
    for c in K["contraste"]:
        print(f"  {'OK ' if c['ok'] else '≠  '} {c['que']}: declara {c['declarado']} · "
              f"calculado {c['calculado']}")
    for d in K["difDT"]:
        print(f"  ⚠  DT {d['prio']} abiertos: hoja STATUS {d['status']} · "
              f"registro DT {d['registro']} (el panel usa el registro)")
    if not K["estadoPropio"]:
        print("  i   El registro no trae columna STATUS: el estado se calculó con la regla "
              "homologada (atrasado = abierto con compromiso vencido)")
    elif K["incoherentes"]:
        print(f"  ⚠  {K['incoherentes']} registros con estado inconsistente respecto de sus fechas")
    else:
        print("  OK  Todos los DT respetan la regla: atrasado = abierto con compromiso vencido")

    if avisos:
        print("\n⚠  AVISOS — revisar antes de publicar:")
        for x in sorted(set(avisos))[:20]:
            print(f"   · {x}")

    print(f"\nEscritos: {destino.name}" + (", index.html (pestaña Talabre)" if inyectado else ""))


if __name__ == "__main__":
    main()
