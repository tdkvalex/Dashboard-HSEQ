#!/usr/bin/env python3
"""
Actualiza el panel de control y los datos del PPT desde el Excel de estatus QAQC.

Uso (cada lunes):
    python3 actualizar.py /ruta/al/Estatus_Resumen_General_QAQC.xlsx [--hoy AAAA-MM-DD]

El corte lo declara el propio Excel («Fecha actualización», hoja Resumen). `--hoy`
es solo el respaldo para cuando esa celda no viene: sin él se usaría la fecha del
día y el proyecto quedaría con un corte distinto al de los otros dos.

Qué hace:
  1. Lee las hojas «BD Caminatas-CTOP» y «BD Detalles Terminación».
  2. Normaliza los estados (tolera variantes de tilde/género: Observada/Observado,
     Rechazada/Rechazado) y AVISA si aparece un estado nuevo no reconocido,
     en lugar de clasificarlo en silencio.
  3. Recalcula todos los indicadores y escribe:
       - estatus_datos.json   (datos consolidados)
       - index.html           (panel autocontenido, datos inyectados)
  4. Imprime un resumen y el comparativo contra la corrida anterior.

Los porcentajes NO se escriben a mano en ninguna parte: se derivan de estos datos.
"""

import json
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

# La apertura de los Excel vive en `qaqc_excel.py`, en la raíz del repositorio:
# descarta los estilos con nombre —que en el REPORTE_GERENCIAL son 11 de sus 12
# MB— y abre en modo read_only, con lo que un corte pasa de ~30 s a ~5 s. Si el
# módulo se copia suelto a otra parte, se cae a la apertura normal de openpyxl y
# todo sigue funcionando, solo que más lento.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
try:
    from qaqc_excel import libro, hoja        # noqa: F401
except ImportError:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("Falta openpyxl.  Instálalo con:  pip install openpyxl")

    def libro(ruta, **kw):
        kw.setdefault("data_only", True)
        p = Path(ruta)
        if not p.exists():
            sys.exit(f"No existe el archivo: {p}")
        return load_workbook(p, **kw)

    def hoja(wb, nombre, ruta):
        if nombre not in wb.sheetnames:
            sys.exit(f"«{Path(ruta).name}» no trae la hoja «{nombre}». "
                     f"Trae: {', '.join(wb.sheetnames[:8])}")
        return wb[nombre]


AQUI = Path(__file__).resolve().parent

AREAS = ["2000", "3000", "4000", "8000"]
AREA_NOMBRES = {
    "2000": "Chancado (Área Seca)",
    "3000": "Molienda–Flotación (Área Húmeda)",
    "4000": "Espesamiento y Relaves",
    "8000": "Salas de Control",
}

# --- Diccionarios de normalización -------------------------------------------
# clave = texto normalizado (sin tilde, minúsculas); valor = estado canónico.
CAMINATA_MAP = {
    "realizada": "Realizada",
    "realizado": "Realizada",
    "proxima a realizar": "Próxima a realizar",
    "proximo a realizar": "Próxima a realizar",
    "programar": "Por programar",
    "por programar": "Por programar",
}

CTOP_MAP = {
    "": "Sin entregar",
    "en revision": "En revisión",
    "revision": "En revisión",
    "observada": "Observada",
    "observado": "Observada",
    "rechazada": "Rechazada",
    "rechazado": "Rechazada",
    "aprobada": "Aprobada",
    "aprobado": "Aprobada",
}
# Una carpeta cuenta como ENTREGADA si ya entró al circuito de revisión.
CTOP_ENTREGADAS = {"En revisión", "Observada", "Rechazada", "Aprobada"}

# Estados de los detalles de terminación, agrupados en 3 categorías de gestión.
DT_MAP = {
    "cerrado": ("Cerrado", "cerrado"),
    "trabajo requerido": ("Trabajo requerido", "abierto"),
    "iniciado": ("Iniciado", "abierto"),
    "trabajo no aceptado": ("Trabajo no aceptado", "abierto"),
    "listo para revision": ("Listo para revisión", "tramite"),
    "listo para cerrar": ("Listo para cerrar", "tramite"),
}

ESPECIALIDAD_MAP = {
    "electrica": "Eléctrica",
    "estructura": "Estructura",
    "instrumentacion y control": "Instrumentación y Control",
    "mecanica": "Mecánica",
    "oo.cc": "Obras Civiles",
    "obras civiles": "Obras Civiles",
    "piping": "Piping",
}
# Orden fijo para el heatmap y la leyenda (etiquetas cortas para el heatmap).
ESPECIALIDADES = ["Piping", "Eléctrica", "Instrumentación y Control",
                  "Mecánica", "Estructura", "Obras Civiles"]
ESP_CORTA = {"Instrumentación y Control": "Instr. y Control"}

avisos = []


def norm(v):
    """Minúsculas, sin tildes, sin espacios sobrantes."""
    if v is None:
        return ""
    s = str(v).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    return " ".join(s.split())


def mapear(valor, tabla, campo):
    """Traduce un valor crudo a su forma canónica; avisa si es desconocido."""
    clave = norm(valor)
    if clave in tabla:
        return tabla[clave]
    avisos.append(f"{campo}: valor no reconocido «{valor}» — revisar clasificación")
    return str(valor).strip() if valor else "Sin dato"


def as_fecha(v):
    if isinstance(v, datetime):
        return v
    if isinstance(v, str) and v.strip():
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(v.strip(), fmt)
            except ValueError:
                pass
    return None


def leer(ruta, respaldo=None):
    wb = libro(ruta)

    # --- fecha de corte, desde la hoja Resumen -------------------------------
    corte = None
    if "Resumen" in wb.sheetnames:
        for fila in wb["Resumen"].iter_rows(min_row=1, max_row=6, values_only=True):
            for i, celda in enumerate(fila):
                if isinstance(celda, str) and "fecha actualiz" in norm(celda):
                    for siguiente in fila[i + 1:]:
                        if isinstance(siguiente, datetime):
                            corte = siguiente
                            break
            if corte:
                break
    if corte is None:
        corte = respaldo or datetime.now()
        avisos.append("No se encontró «Fecha actualización» en la hoja Resumen; "
                      f"se usa {corte:%d-%m-%Y} para calcular vencidos"
                      + ("" if respaldo else " (la fecha de hoy: pásala con --hoy "
                         "si estás reprocesando un corte pasado)") + ".")

    # --- caminatas y carpetas TOP -------------------------------------------
    caminatas = []
    for fila in hoja(wb, "BD Caminatas-CTOP", ruta).iter_rows(min_row=2, values_only=True):
        if fila[0] is None:
            continue
        caminatas.append({
            "clasificacion": (fila[1] or "").strip(),
            "area": str(fila[2]).strip(),
            "subsistema": str(fila[3]).strip(),
            "c80": mapear(fila[4], CAMINATA_MAP, "Caminata 80%"),
            "c100": mapear(fila[5], CAMINATA_MAP, "Caminata 100%"),
            "ctop": mapear(fila[6], CTOP_MAP, "Estatus CTOP"),
        })

    # --- detalles de terminación --------------------------------------------
    detalles = []
    for fila in hoja(wb, "BD Detalles Terminación", ruta).iter_rows(min_row=2, values_only=True):
        if all(v is None for v in fila):
            continue
        estado, grupo = mapear(fila[8], DT_MAP, "Estatus detalle")\
            if norm(fila[8]) in DT_MAP else (str(fila[8]).strip(), "abierto")
        if norm(fila[8]) not in DT_MAP:
            avisos.append(f"Estatus detalle no reconocido «{fila[8]}» — "
                          f"contabilizado como ABIERTO por precaución")
        vence = as_fecha(fila[9])
        detalles.append({
            "area": str(fila[1]).strip(),
            "tipo": (fila[4] or "").strip(),
            "especialidad": mapear(fila[5], ESPECIALIDAD_MAP, "Especialidad"),
            "estado": estado,
            "grupo": grupo,
            "vencido": bool(grupo == "abierto" and vence and vence < corte),
            "caminata": (fila[13] or "").strip(),
        })

    return corte, caminatas, detalles


def resumir(filas):
    """Agrega un conjunto de detalles en el resumen de gestión."""
    return {
        "total": len(filas),
        "cerrados": sum(1 for d in filas if d["grupo"] == "cerrado"),
        "tramite": sum(1 for d in filas if d["grupo"] == "tramite"),
        "abiertos": sum(1 for d in filas if d["grupo"] == "abierto"),
        "vencidos": sum(1 for d in filas if d["vencido"]),
    }


def construir(corte, caminatas, detalles):
    p1 = [d for d in detalles if d["tipo"] == "P1"]

    cam_area = {}
    for a in AREAS:
        sub = [c for c in caminatas if c["area"] == a]
        cam_area[a] = {
            "subs": len(sub),
            "c80": sum(1 for c in sub if c["c80"] == "Realizada"),
            "c100": sum(1 for c in sub if c["c100"] == "Realizada"),
            "prox": sum(1 for c in sub if c["c100"] == "Próxima a realizar"),
            "prog": sum(1 for c in sub if c["c100"] == "Por programar"),
        }

    ctop_area = {}
    for a in AREAS:
        sub = [c for c in caminatas if c["area"] == a]
        cnt = Counter(c["ctop"] for c in sub)
        ctop_area[a] = {
            "total": len(sub),
            "rev": cnt.get("En revisión", 0),
            "obs": cnt.get("Observada", 0),
            "rech": cnt.get("Rechazada", 0),
            "aprob": cnt.get("Aprobada", 0),
        }

    ctop_global = Counter(c["ctop"] for c in caminatas)
    entregadas = sum(v for k, v in ctop_global.items() if k in CTOP_ENTREGADAS)

    heat = []
    for e in ESPECIALIDADES:
        heat.append([sum(1 for d in p1 if d["especialidad"] == e
                         and d["area"] == a and d["grupo"] == "abierto")
                     for a in AREAS])

    datos = {
        "meta": {
            "corte": corte.strftime("%Y-%m-%d"),
            "corte_texto": corte.strftime("%d-%m-%Y"),
            "generado": datetime.now().strftime("%Y-%m-%d %H:%M"),
        },
        "areas": AREAS,
        "areaNames": AREA_NOMBRES,
        "cam": {
            "global": {
                "subs": len(caminatas),
                "c80": sum(1 for c in caminatas if c["c80"] == "Realizada"),
                "c100": sum(1 for c in caminatas if c["c100"] == "Realizada"),
                "prox": sum(1 for c in caminatas if c["c100"] == "Próxima a realizar"),
                "prog": sum(1 for c in caminatas if c["c100"] == "Por programar"),
            },
            "byArea": cam_area,
            "pend": [[c["area"], c["subsistema"], c["c100"]]
                     for c in caminatas if c["c100"] != "Realizada"],
        },
        "ctop": {
            "global": {
                "total": len(caminatas),
                "entregadas": entregadas,
                "rev": ctop_global.get("En revisión", 0),
                "obs": ctop_global.get("Observada", 0),
                "rech": ctop_global.get("Rechazada", 0),
                "aprob": ctop_global.get("Aprobada", 0),
                "pend": ctop_global.get("Sin entregar", 0),
            },
            "byArea": ctop_area,
            "detalle": [[c["area"], c["subsistema"], c["ctop"]]
                        for c in caminatas if c["ctop"] in CTOP_ENTREGADAS],
        },
        "dt": {
            "global": resumir(detalles),
            "tipos": {t: resumir([d for d in detalles if d["tipo"] == t])
                      for t in ("P1", "P2", "P3")},
            "p1Esp": {e: resumir([d for d in p1 if d["especialidad"] == e])
                      for e in ESPECIALIDADES},
            "p1Area": {a: resumir([d for d in p1 if d["area"] == a]) for a in AREAS},
            "hm": {"esp": [ESP_CORTA.get(e, e) for e in ESPECIALIDADES], "vals": heat},
            "p1AbiertosPorOrigen": dict(Counter(
                d["caminata"] for d in p1 if d["grupo"] == "abierto")),
        },
    }
    return datos


def semaforo(datos):
    """Clasifica cada área. Devuelve {area: (clase, etiqueta)}."""
    out = {}
    for a in datos["areas"]:
        c = datos["cam"]["byArea"][a]
        p = datos["dt"]["p1Area"][a]
        t = datos["ctop"]["byArea"][a]
        pct_c100 = 100 * c["c100"] / c["subs"] if c["subs"] else 0
        pct_p1 = 100 * p["cerrados"] / p["total"] if p["total"] else 0
        entregadas = t["rev"] + t["obs"] + t["rech"] + t["aprob"]
        if pct_c100 < 60 or pct_p1 < 50:
            out[a] = ("crit", "Crítico")
        elif pct_c100 >= 100 and pct_p1 >= 90 and entregadas == t["total"]:
            out[a] = ("good", "Al día")
        elif pct_c100 >= 90 and pct_p1 >= 90:
            out[a] = ("good", "Al día")
        else:
            out[a] = ("warn", "Atención")
    return out


def comparar(nuevo, anterior):
    # Solo compara si el archivo previo tiene el mismo esquema que este script
    # genera (una corrida anterior de actualizar.py).
    if not anterior or "cam" not in anterior or "meta" not in anterior:
        return []
    lineas = []

    def delta(etiqueta, a, b, sufijo=""):
        d = b - a
        if d:
            lineas.append(f"  {etiqueta}: {a} → {b} ({d:+d}{sufijo})")

    delta("Caminatas 100% realizadas",
          anterior["cam"]["global"]["c100"], nuevo["cam"]["global"]["c100"])
    delta("Carpetas TOP entregadas",
          anterior["ctop"]["global"]["entregadas"], nuevo["ctop"]["global"]["entregadas"])
    delta("Carpetas TOP rechazadas",
          anterior["ctop"]["global"]["rech"], nuevo["ctop"]["global"]["rech"])
    delta("Carpetas TOP aprobadas",
          anterior["ctop"]["global"]["aprob"], nuevo["ctop"]["global"]["aprob"])
    delta("P1 cerrados",
          anterior["dt"]["tipos"]["P1"]["cerrados"], nuevo["dt"]["tipos"]["P1"]["cerrados"])
    delta("P1 abiertos",
          anterior["dt"]["tipos"]["P1"]["abiertos"], nuevo["dt"]["tipos"]["P1"]["abiertos"])
    delta("P1 vencidos",
          anterior["dt"]["tipos"]["P1"]["vencidos"], nuevo["dt"]["tipos"]["P1"]["vencidos"])
    return lineas


def actualizar_historial(datos):
    """Mantiene historial.json: una entrada por fecha de corte.

    Permite mostrar la variación semana a semana en el panel y el PPT.
    Si se vuelve a procesar un corte ya registrado, se reemplaza esa entrada
    (para poder corregir una carga sin duplicar la semana).
    """
    ruta = AQUI / "historial.json"
    historia = []
    if ruta.exists():
        try:
            historia = json.loads(ruta.read_text(encoding="utf-8"))
        except Exception:
            avisos.append("historial.json ilegible; se recrea desde este corte.")
            historia = []

    c, t, p1 = datos["cam"]["global"], datos["ctop"]["global"], datos["dt"]["tipos"]["P1"]
    entrada = {
        "corte": datos["meta"]["corte"],
        "corte_texto": datos["meta"]["corte_texto"],
        "cam80": c["c80"], "cam100": c["c100"], "subs": c["subs"],
        "camProx": c["prox"], "camProg": c["prog"],
        "topEntregadas": t["entregadas"], "topRev": t["rev"],
        "topObs": t["obs"], "topRech": t["rech"], "topAprob": t["aprob"],
        "p1Total": p1["total"], "p1Cerrados": p1["cerrados"],
        "p1Abiertos": p1["abiertos"], "p1Vencidos": p1["vencidos"],
    }
    historia = [h for h in historia if h.get("corte") != entrada["corte"]]
    historia.append(entrada)
    historia.sort(key=lambda h: h["corte"])
    ruta.write_text(json.dumps(historia, ensure_ascii=False, indent=1), encoding="utf-8")
    return historia


MARCA_INI = "<!-- === DATOS:INICIO ==="
MARCA_FIN = "<!-- === DATOS:FIN === -->"


def inyectar_en_html(ruta_html, datos):
    """Reemplaza el bloque de datos dentro de index.html.

    Se inyecta en el propio HTML (en vez de un .js aparte) para que el panel
    siga siendo UN archivo portable: se puede enviar por correo tal cual.
    """
    if not ruta_html.exists():
        avisos.append(f"No se encontró {ruta_html.name}; solo se escribió el JSON.")
        return
    html = ruta_html.read_text(encoding="utf-8")
    i, f = html.find(MARCA_INI), html.find(MARCA_FIN)
    if i == -1 or f == -1 or f < i:
        avisos.append(f"No se encontraron las marcas DATOS en {ruta_html.name}; "
                      f"el panel quedó con los datos anteriores.")
        return
    bloque = (
        MARCA_INI + " (bloque generado por actualizar.py — no editar a mano) -->\n"
        "<script>\nconst D = "
        + json.dumps(datos, ensure_ascii=False, indent=1)
        + ";\n</script>\n"
    )
    ruta_html.write_text(html[:i] + bloque + html[f:], encoding="utf-8")


def main():
    # El corte de Arqueros lo declara el propio Excel («Fecha actualización» de
    # la hoja Resumen), que es la fuente correcta. --hoy solo existe para el
    # respaldo: si esa celda no viene, antes se usaba la fecha del día en
    # silencio y el proyecto quedaba con un corte distinto al de los otros dos.
    argv = sys.argv[1:]
    respaldo = None
    if "--hoy" in argv:
        i = argv.index("--hoy")
        if i + 1 >= len(argv):
            sys.exit("--hoy necesita una fecha AAAA-MM-DD")
        try:
            respaldo = datetime.strptime(argv[i + 1], "%Y-%m-%d")
        except ValueError:
            sys.exit("--hoy debe ser AAAA-MM-DD")
        argv = argv[:i] + argv[i + 2:]
    if not argv:
        sys.exit(__doc__)
    ruta = Path(argv[0]).expanduser()
    if not ruta.exists():
        sys.exit(f"No existe el archivo: {ruta}")

    destino_json = AQUI / "estatus_datos.json"
    anterior = None
    if destino_json.exists():
        try:
            anterior = json.loads(destino_json.read_text(encoding="utf-8"))
        except Exception:
            anterior = None

    corte, caminatas, detalles = leer(ruta, respaldo)
    datos = construir(corte, caminatas, detalles)
    datos["semaforo"] = semaforo(datos)

    historia = actualizar_historial(datos)
    # La entrada previa (si existe) alimenta la variación semanal del panel.
    previas = [h for h in historia if h["corte"] < datos["meta"]["corte"]]
    datos["previo"] = previas[-1] if previas else None

    destino_json.write_text(
        json.dumps(datos, ensure_ascii=False, indent=1), encoding="utf-8")
    inyectar_en_html(AQUI / "index.html", datos)

    g, c, t, p1 = (datos["dt"]["global"], datos["cam"]["global"],
                   datos["ctop"]["global"], datos["dt"]["tipos"]["P1"])
    pc = lambda a, b: f"{100*a/b:.1f}%".replace(".", ",") if b else "—"

    print(f"\nCorte: {datos['meta']['corte_texto']}")
    print(f"  Caminatas 80%       {c['c80']}/{c['subs']}  {pc(c['c80'], c['subs'])}")
    print(f"  Caminatas 100%      {c['c100']}/{c['subs']}  {pc(c['c100'], c['subs'])}")
    print(f"  Carpetas TOP entr.  {t['entregadas']}/{t['total']}  "
          f"{pc(t['entregadas'], t['total'])}   "
          f"(aprobadas {t['aprob']} · rechazadas {t['rech']} · en revisión {t['rev']})")
    print(f"  Cierre P1           {p1['cerrados']}/{p1['total']}  "
          f"{pc(p1['cerrados'], p1['total'])}   "
          f"(abiertos {p1['abiertos']} · vencidos {p1['vencidos']})")
    print(f"  Detalles totales    {g['total']}  "
          f"(cerrados {g['cerrados']} · trámite {g['tramite']} · abiertos {g['abiertos']})")

    cambios = comparar(datos, anterior)
    if cambios:
        print("\nCambios respecto de la corrida anterior "
              f"(corte {anterior['meta']['corte_texto']}):")
        print("\n".join(cambios))

    if avisos:
        print("\n⚠  AVISOS — revisar antes de publicar:")
        for a in sorted(set(avisos)):
            print(f"  · {a}")

    print(f"\nEscritos: {destino_json.name}, index.html")
    print("index.html quedó actualizado y sigue siendo un archivo único portable.")


if __name__ == "__main__":
    main()
