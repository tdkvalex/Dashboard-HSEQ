#!/usr/bin/env python3
"""Protocolos · P2342 Arqueros (cliente MASA) — carga de las matrices de disciplina.

Lee las matrices que entrega el proyecto —una por disciplina— y escribe el bloque
`P2342` del dashboard de Protocolos, su historial por nodo y el punto semanal del
KPI. No toca los otros dos proyectos.

Por qué este script existe y no se leen las cifras que el propio libro muestra
--------------------------------------------------------------------------------
Cada matriz trae una hoja **`KPI-BSMT`** —la única que se usa, nunca `KPI-MASA`,
que es el alcance del cliente— con el recuento por estado. Esa hoja **no se lee**:
se **recalcula** aplicando sus mismas fórmulas sobre la matriz. El motivo es que
sus valores vienen guardados de la última vez que Excel recalculó, y en varias
matrices eso fue hace tiempo:

* En **Obras Civiles** las fórmulas `UNIQUE`/`FILTER` quedaron guardadas como `1`
  en cada columna —el desborde dinámico colapsado—, así que la hoja declara
  **16 cerrados** donde hay **4.748**. Publicado tal cual, el proyecto aparecería
  desplomándose de un corte al otro sin que nada hubiera pasado.
* En Estructuras, Piping, Comunicación e Instrumentación la caché está sólo
  desfasada (diferencias de 1 a 25 registros).

Donde la caché está sana, este recálculo da **exactamente** su mismo número; eso
es lo que lo valida. `--verificar-cache` imprime la comparación columna por columna.

Cómo cuenta el libro (y por eso, cómo contamos aquí)
-----------------------------------------------------
La unidad **no es la misma para todos los estados**, y así está en el original:

| Estado | Fórmula del libro | Unidad |
|---|---|---|
| `C`, `AP`, `AE` | `COUNTA(UNIQUE(FILTER(PROTOCOLO N°; ESTATUS="x")))` | protocolo distinto |
| `S`, `P`, `N`   | `COUNTIF(ESTATUS;"x")` | celda elemento×protocolo |

El «universo» que arma la hoja suma las dos cosas. Es una inconsistencia del
formato de origen, no un error de lectura: se replica tal cual para que el
dashboard muestre lo mismo que MASA y Besalco revisan en la reunión.

Además, cada columna del KPI suma la matriz de protocolos **y** la de precom del
mismo libro (`Matriz PreCom`, `Matriz de Precom`, `Matriz Precom-EEII`), filtrando
por la sigla que va en la fila 2 sobre cada columna de ESTATUS.

Uso
---
    python3 protocolos_masa.py --matrices <carpeta con las matrices>
    python3 protocolos_masa.py --matrices <carpeta> --corte 2026-08-17
    python3 protocolos_masa.py --matrices <carpeta> --verificar-cache
    python3 protocolos_masa.py --matrices <carpeta> --seco   # no escribe nada
"""

import argparse
import glob
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
try:
    from qaqc_excel import libro, hojas          # noqa: F401
except ImportError:                              # copiado suelto a otra parte
    import zipfile
    from xml.etree import ElementTree
    from openpyxl import load_workbook

    def libro(ruta, read_only=True, **kw):
        kw.setdefault("data_only", True)
        return load_workbook(ruta, read_only=read_only, **kw)

    def hojas(ruta):
        try:
            with zipfile.ZipFile(ruta) as z:
                raiz = ElementTree.fromstring(z.read("xl/workbook.xml"))
            ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
            return [h.get("name") for h in raiz.iter(f"{ns}sheet") if h.get("name")]
        except Exception:
            return []

from openpyxl import load_workbook as _lw_raw
from openpyxl.utils import range_boundaries

PANEL = Path(__file__).resolve().parent.parent / "suite_qaqc" / "modulos" / "protocolos.html"

ESTADOS = ("S", "C", "P", "AP", "AE", "N")
# Los que el libro cuenta por protocolo distinto, no por celda.
POR_PROTOCOLO = ("C", "AP", "AE")

# Las matrices se reconocen por contenido: los nombres de archivo cambian entre
# cortes. El título de «Matriz de Protocolos» trae el código del documento,
# `BSMT-P23/42-85-<SIGLA>-MTZ-01`, que identifica disciplina y proyecto a la vez.
CODIGO = re.compile(r"BSMT-P(\d\d)/(\d\d)-\d+-([A-ZÁÉÍÓÚ]+)-MTZ", re.I)
POR_CODIGO = {"OOCC": "OOCC", "ESTR": "ESTR", "MECA": "MECA", "PIPN": "PIPN",
              "ELEC": "ELEC", "INS": "INST", "INST": "INST", "COM": "COMU",
              "COMU": "COMU"}

# Respaldo por si algún corte llega sin el código en el título.
TITULOS = {
    "OOCC": ("OO.CC", "OBRAS CIVILES"),
    "ESTR": ("ESTRUCTURA",),
    "MECA": ("EQUIPOS MEC", "MECÁNIC", "MECANIC"),
    "PIPN": ("PIPING",),
    "ELEC": ("ELÉCTRIC", "ELECTRIC"),
    "INST": ("INSTRUMENTA",),
    "COMU": ("COMUNICA",),
}

PROYECTO = "2342"

# Nodo del dashboard → (matriz, siglas que suma esa columna del KPI-BSMT).
MAPA = {
    "TOPO-OOCC-2342":   ("OOCC", ["TOPO"]),
    "TOPO-MECA-2342":   ("MECA", ["TOPO"]),
    "TOPO-PIPN-2342":   ("PIPN", ["TOPO"]),
    "TOPO-ESTR-2342":   ("ESTR", ["TOPO"]),
    "TOPO-ELEC-2342":   ("ELEC", ["TOPO"]),
    "OOCC-2342":        ("OOCC", ["OOCC"]),
    "ESTR-2342":        ("ESTR", ["ESTR"]),
    "MECA-2342":        ("MECA", ["MECA"]),
    "PIPN-2342":        ("PIPN", ["PIPN"]),
    "ELEC-A-2342":      ("ELEC", ["EEII"]),
    "INST-A-2342":      ("INST", ["INST", "EEII"]),
    "COMU-2342":        ("COMU", ["EEII"]),
    "PCOM-ELEC-2342":   ("ELEC", ["PCOM"]),
    "PCOM-INST-2342":   ("INST", ["PCOM"]),
    "PCOM-MECA-2342":   ("MECA", ["PCOM"]),
    "PCOM-COMU-2342":   ("COMU", ["PCOM", "PREC"]),
}

# Nodos sin matriz propia desde el corte 29-06-2026: la matriz de la línea BT/MT
# dejó de llegar. Conservan su último valor conocido y NO reciben punto nuevo de
# historial —inventarle continuidad a un dato que nadie actualizó sería peor que
# dejarlo quieto—. El script lo declara en AVISOS en cada corrida.
CONGELADOS = {
    "TOPO-ELECBT-2342": dict(S=98, C=0, P=0, AP=0, AE=0),
    "ELEC-BT-2342":     dict(S=65, C=0, P=0, AP=0, AE=7),
}

# Estructura del proyecto en el panel: (id, nombre, [hijos]) — los hijos son
# (id, nombre, etiqueta de matriz).
ESTRUCTURA = [
    ("TOPO-2342", "Topografía", [
        ("TOPO-OOCC-2342",   "Topo en Obras Civiles",   "MTZ-OOCC"),
        ("TOPO-MECA-2342",   "Topo en Mecánica",        "MTZ-MECA"),
        ("TOPO-PIPN-2342",   "Topo en Piping",          "MTZ-PIPN"),
        ("TOPO-ESTR-2342",   "Topo en Estructuras",     "MTZ-ESTR"),
        ("TOPO-ELEC-2342",   "Topo en Eléctrica",       "MTZ-ELEC"),
        ("TOPO-ELECBT-2342", "Topo en Eléctrica BT/MT", "MTZ-BTMT"),
    ]),
    ("OOCC-2342", "Obras Civiles", None),
    ("ESTR-2342", "Estructuras",   None),
    ("MECA-2342", "Mecánica",      None),
    ("PIPN-2342", "Piping",        None),
    ("ELEC-2342", "Eléctrico", [
        ("ELEC-A-2342",  "Eléctrica (Prot. EEII)", "MTZ-ELEC"),
        ("ELEC-BT-2342", "Eléctrica Línea BT/MT",  "MTZ-BTMT"),
    ]),
    ("INST-2342", "Instrumentación", [
        ("INST-A-2342", "Instrumentación", "MTZ-INST"),
        ("COMU-2342",   "Comunicaciones",  "MTZ-COMU"),
    ]),
    ("PCOM-2342", "Precomisionamiento", [
        ("PCOM-ELEC-2342", "Precom Eléctrica",       "PRECOM-EEII"),
        ("PCOM-INST-2342", "Precom Instrumentación", "PRECOM-INST"),
        ("PCOM-MECA-2342", "Precom Mecánica",        "PRECOM-MECA"),
        ("PCOM-COMU-2342", "Precom Comunicaciones",  "PRECOM-COMU"),
    ]),
]

# Disciplina del KPI_HISTORY → nodos que la componen. Comunicaciones va dentro de
# Instrumentación, igual que en el árbol del panel.
DISCIPLINAS = {
    "Topografía":         ["TOPO-OOCC-2342", "TOPO-MECA-2342", "TOPO-PIPN-2342",
                           "TOPO-ESTR-2342", "TOPO-ELEC-2342", "TOPO-ELECBT-2342"],
    "Obras_Civiles":      ["OOCC-2342"],
    "Estructuras":        ["ESTR-2342"],
    "Mecánica":           ["MECA-2342"],
    "Piping":             ["PIPN-2342"],
    "Eléctrico":          ["ELEC-A-2342", "ELEC-BT-2342"],
    "Instrumentación":    ["INST-A-2342", "COMU-2342"],
    "Precomisionamiento": ["PCOM-ELEC-2342", "PCOM-INST-2342",
                           "PCOM-MECA-2342", "PCOM-COMU-2342"],
}

AVISOS = []


def aviso(t):
    AVISOS.append(t)


def pct2(x):
    """Dos decimales, redondeando como JS. `round()` de Python es bancario
    (`round(0.125,2)` = 0.12) y desalineaba panel, JSON y PPT. Dos decimales
    porque es el formato que ya usan `KPI_HISTORY` y `PPT_PREV_DISC`: con uno
    solo, los KPI de este proyecto —todos por debajo de 1%— se aplastarían
    contra 0,0 y la variación semanal dejaría de verse."""
    import math
    return math.floor(x * 100 + 0.5) / 100


def kpi_pct(d):
    """(P + AP) / (P + AP + AE + C), en porcentaje. Misma fórmula que el libro
    y que `kpiV()` del panel."""
    den = d["P"] + d["AP"] + d["AE"] + d["C"]
    return (d["P"] + d["AP"]) / den * 100 if den else 0.0


# ── lectura ───────────────────────────────────────────────────────────────

def _matrices_del_libro(ruta):
    """Hojas de matriz (protocolos y precom) con el rango de su tabla.

    Abrir el libro sin `data_only` es lo único que da los rangos de las tablas,
    y es también la lectura que más se atraganta: los libros con caché de tabla
    dinámica —el log de NC del cliente, por ejemplo— la revientan. Como el
    auditor pasa por todos los archivos de la carpeta, aquí se falla en silencio
    y el archivo simplemente no cuenta como matriz.
    """
    try:
        wb = _lw_raw(ruta, data_only=False)
    except Exception:
        return []
    out = []
    for sn in wb.sheetnames:
        if not sn.lower().startswith("matriz"):
            continue
        ws = wb[sn]
        if not ws.tables:
            continue
        out.append((sn, list(ws.tables.values())[0].ref))
    wb.close()
    return out


def _titulo_y_fecha(wb):
    """Título de la matriz y la fecha que ella misma declara."""
    ws = wb["Matriz de Protocolos"]
    titulo, fecha = "", None
    for i, fila in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 11:
            break
        for j, v in enumerate(fila[:10]):
            if not isinstance(v, str):
                continue
            t = v.strip().upper()
            if t.startswith("MATRIZ DE PROTOCOLOS") and len(t) > len(titulo):
                titulo = t
            if "FECHA ACTUALIZ" in t:
                for x in fila[j:j + 6]:
                    if hasattr(x, "strftime"):
                        fecha = x.strftime("%Y-%m-%d")
    return titulo, fecha


def leer_matriz(ruta):
    """→ (disciplina, fecha, {sigla: {S,C,P,AP,AE,N}}, {sigla: caché del libro})

    Devuelve `(None, …)` si el archivo no es una matriz de protocolos. Se
    descarta leyendo los nombres de hoja del ZIP, sin abrir el libro: por aquí
    pasan también los Excel de los otros módulos.
    """
    if "Matriz de Protocolos" not in hojas(ruta):
        return None, None, {}, {}
    rangos = _matrices_del_libro(ruta)
    if not rangos:
        return None, None, {}, {}

    wb = libro(ruta)
    if "Matriz de Protocolos" not in wb.sheetnames:
        wb.close()
        return None, None, {}, {}
    titulo, fecha = _titulo_y_fecha(wb)

    disc = None
    m = CODIGO.search(titulo)
    if m:
        if m.group(1) + m.group(2) != PROYECTO:
            aviso(f"«{Path(ruta).name}» declara el proyecto P{m.group(1)}{m.group(2)}, "
                  f"no P{PROYECTO}: no se carga.")
            wb.close()
            return None, None, {}, {}
        disc = POR_CODIGO.get(m.group(3).upper())
    if not disc:
        for d, claves in TITULOS.items():
            if any(k in titulo for k in claves):
                disc = d
                break

    calc = {}
    for hoja_nom, ref in rangos:
        c1, r1, c2, r2 = range_boundaries(ref)
        filas = list(wb[hoja_nom].iter_rows(values_only=True))
        if len(filas) <= r1:
            continue
        cab, siglas = filas[r1 - 1], filas[1]
        cols = [i for i, v in enumerate(cab)
                if isinstance(v, str) and v.strip().upper().startswith("ESTATUS")
                and c1 - 1 <= i <= c2 - 1]
        for i in cols:
            sig = str(siglas[i]).strip() if i < len(siglas) and siglas[i] else None
            if not sig:
                continue
            d = calc.setdefault(sig, {k: 0 for k in ESTADOS})
            unicos = {k: set() for k in POR_PROTOCOLO}
            for f in filas[r1:r2]:
                v = str(f[i]).strip().upper() if i < len(f) and f[i] is not None else ""
                if v not in ESTADOS:
                    continue
                if v in POR_PROTOCOLO:
                    p = f[i - 1]
                    unicos[v].add(str(p).strip() if p is not None else "")
                else:
                    d[v] += 1
            for k in POR_PROTOCOLO:
                d[k] += len(unicos[k])

    cache = _leer_kpi_bsmt(wb, ruta)
    wb.close()
    return disc, fecha, calc, cache


def _leer_kpi_bsmt(wb, ruta):
    """Lo que la hoja KPI-BSMT trae guardado. Sólo para contrastar."""
    if "KPI-BSMT" not in wb.sheetnames:
        aviso(f"«{Path(ruta).name}» no trae la hoja KPI-BSMT. "
              f"No se usa KPI-MASA como reemplazo: pide la matriz correcta.")
        return {}
    filas = list(wb["KPI-BSMT"].iter_rows(values_only=True))
    cab = next((f for f in filas
                if f and any(isinstance(x, str) and x.strip() == "Estatus" for x in f[:4])), None)
    if not cab:
        return {}
    cols = {v.strip(): i for i, v in enumerate(cab)
            if isinstance(v, str) and v.strip() not in ("Estatus", "TOTAL", "KPI", "")}
    out = {}
    for f in filas:
        et = next((str(x).strip() for x in f[:3]
                   if isinstance(x, str) and x.strip().startswith("(")), None)
        if not et:
            continue
        k = et.split(")")[0].lstrip("(")
        if k in ESTADOS:
            for sig, i in cols.items():
                v = f[i] if i < len(f) else None
                out.setdefault(sig, {})[k] = v if isinstance(v, (int, float)) else 0
    return out


# ── escritura en el panel ─────────────────────────────────────────────────

def bloque_proyecto(datos, corte_iso, fechas):
    """El literal `P2342:{…}` completo."""
    d = datetime.strptime(corte_iso, "%Y-%m-%d")
    MES = ["ene", "feb", "mar", "abr", "may", "jun",
           "jul", "ago", "sep", "oct", "nov", "dic"]
    visible = f"{d.day:02d}-{MES[d.month - 1]}-{d.year}"

    def cifras(nid):
        v = datos[nid]
        return (f"S:{v['S']}, C:{v['C']}, P:{v['P']}, AP:{v['AP']}, AE:{v['AE']}")

    L = []
    L.append("  P2342:{")
    L.append(f"    name:'P2342 — Proyecto Arqueros', client:'MASA', updated:'{visible}',")
    L.append(f"    lastUpload:{{all:'{corte_iso}T00:00:00'}},")
    L.append("    status:'active', contract:'', startDate:'', sheetName:'KPI-BSMT', hasGoformz:false,")
    L.append("    rows:[")
    for nid, nombre, hijos in ESTRUCTURA:
        if hijos is None:
            L.append(f"      {{ id:'{nid}', name:'{nombre}', {cifras(nid)} }},")
        else:
            L.append(f"      {{ id:'{nid}', name:'{nombre}', children:[")
            for cid, cnom, tag in hijos:
                cong = "  // sin matriz desde 29-06-2026" if cid in CONGELADOS else ""
                L.append(f"        {{id:'{cid}', name:'{cnom}', matrixTag:'{tag}', "
                         f"{cifras(cid)} }},{cong}")
            L.append("      ]},")
    L.append("    ]")
    L.append("  }")
    return "\n".join(L)


def inyectar(html, datos, corte_iso, fechas):
    # 1 · bloque del proyecto
    i = html.find("\n  P2342:{")
    if i < 0:
        sys.exit("No encuentro el bloque «P2342:{» en el panel.")
    fin = html.find("\n  }\n};", i)
    if fin < 0:
        sys.exit("No encuentro el cierre del bloque P2342.")
    html = html[:i + 1] + bloque_proyecto(datos, corte_iso, fechas) + html[fin + len("\n  }"):]

    # 2 · historial por nodo — los congelados no reciben punto nuevo
    for nid, v in datos.items():
        if nid in CONGELADOS:
            continue
        u = v["S"] + v["C"] + v["P"] + v["AP"] + v["AE"]
        ent = (f",{{date:'{corte_iso}',loadedBy:'sistema',source:'Excel matrices',"
               f"data:{{S:{v['S']},C:{v['C']},P:{v['P']},AP:{v['AP']},AE:{v['AE']}}},"
               f"kpiPct:{kpi_pct(v):.2f},universe:{u}}}")
        pat = re.compile(r"('" + re.escape(nid) + r"':\[)(.*?)(\],?\n)", re.S)
        m = pat.search(html)
        if not m:
            # nodo nuevo: se abre su historial con este corte
            ancla = "  P2342: {"
            k = html.find(ancla, html.find("let NODE_HISTORY"))
            if k < 0:
                aviso(f"No pude abrir historial para el nodo nuevo {nid}.")
                continue
            nuevo = (f"\n    '{nid}':[{ent[1:]}],")
            html = html[:k + len(ancla)] + nuevo + html[k + len(ancla):]
            aviso(f"{nid}: nodo nuevo, historial abierto en {corte_iso}.")
            continue
        if f"date:'{corte_iso}'" in m.group(2):
            aviso(f"{nid}: ya había un punto en {corte_iso}; se reemplaza.")
            cuerpo = re.sub(r",?\{date:'" + corte_iso + r"'.*?universe:\d+\}", "", m.group(2))
            html = html[:m.start()] + m.group(1) + cuerpo + ent + m.group(3) + html[m.end():]
        else:
            html = html[:m.start()] + m.group(1) + m.group(2) + ent + m.group(3) + html[m.end():]
    return html


def _kpi_proyectos(html):
    """KPI de cada proyecto y el corporativo, leídos del panel ya escrito.

    Se suman las hojas del árbol —los nodos padre no llevan cifras propias— con
    la misma fórmula de `kpiV()`. Calcularlo aquí, y no arrastrarlo del punto
    anterior, es lo que evita que el corporativo se quede pegado cuando cambia
    un solo proyecto.
    """
    i = html.find("const PROJECTS = {")
    fin = html.find("\n};", i)
    blk = html[i:fin]
    pos = sorted((blk.find(p + ":{"), p) for p in ("P2416", "P2407", "P2342") if blk.find(p + ":{") >= 0)
    out, corp = {}, {k: 0 for k in ESTADOS}
    for n, (a, pid) in enumerate(pos):
        b = pos[n + 1][0] if n + 1 < len(pos) else len(blk)
        seg = blk[a:b]
        if re.search(r"status:'(?!active)", seg):      # proyecto cerrado
            continue
        t = {k: 0 for k in ESTADOS}
        for h in re.finditer(r"S:(\d+),\s*C:(\d+),\s*P:(\d+),\s*AP:(\d+),\s*AE:(\d+)", seg):
            for k, v in zip(("S", "C", "P", "AP", "AE"), h.groups()):
                t[k] += int(v)
        for k in ESTADOS:
            corp[k] += t[k]
        out[pid] = pct2(kpi_pct(t))
    out["CORP"] = pct2(kpi_pct(corp))
    return out


def inyectar_kpi(html, datos, corte_iso):
    """Punto semanal del KPI. Arqueros se recalcula; los otros dos proyectos
    conservan su desglose por disciplina hasta que entreguen sus matrices."""
    disc = {}
    for nombre, nodos in DISCIPLINAS.items():
        t = {k: 0 for k in ESTADOS}
        for n in nodos:
            for k in ESTADOS:
                t[k] += datos[n].get(k, 0)
        disc[nombre] = pct2(kpi_pct(t))

    i = html.find("let KPI_HISTORY")
    fin = html.find("\n];", i)
    bloque = html[i:fin]

    # Desglose de los otros dos proyectos: se copia del último punto, para que la
    # variación semanal por disciplina no se corte cuando sólo cambia Arqueros.
    otros = ""
    ult = list(re.finditer(r"\{ ?date:'(\d{4}-\d\d-\d\d)'", bloque))
    if ult:
        cola = bloque[ult[-1].start():]
        for pid in ("P2407", "P2416"):
            m = re.search(pid + r":\{[^{}]*\}", cola[cola.find("disc:"):] if "disc:" in cola else "")
            if m:
                otros += "," + m.group(0)
        if ult[-1].group(1) == corte_iso:
            aviso(f"KPI_HISTORY ya tenía {corte_iso}; se reemplaza el punto.")
            bloque = bloque[:ult[-1].start()].rstrip().rstrip(",")
            html = html[:i] + bloque + html[fin:]
            fin = html.find("\n];", i)

    k = _kpi_proyectos(html)
    d_txt = ",".join(f"{a}:{b}" for a, b in disc.items())
    ent = (f",\n  {{date:'{corte_iso}', kpi:{{P2342:{k['P2342']}, P2407:{k['P2407']}, "
           f"P2416:{k['P2416']}, CORP:{k['CORP']}}}, "
           f"disc:{{P2342:{{{d_txt}}}{otros}}}}}")
    html = html[:fin] + ent + html[fin:]
    return html, k, disc


# ── principal ─────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--matrices", required=True,
                    help="carpeta con las matrices de disciplina (.xlsx)")
    ap.add_argument("--corte", default=None,
                    help="fecha del corte AAAA-MM-DD (por defecto, el lunes de esta semana)")
    ap.add_argument("--panel", default=str(PANEL))
    ap.add_argument("--verificar-cache", action="store_true",
                    help="compara el recálculo contra lo guardado en KPI-BSMT")
    ap.add_argument("--seco", action="store_true", help="no escribe el panel")
    a = ap.parse_args()

    if a.corte:
        corte = a.corte
    else:
        h = date.today()
        corte = (h.toordinal() - h.weekday())
        corte = date.fromordinal(corte).isoformat()

    rutas = sorted(glob.glob(os.path.join(a.matrices, "*.xls*")))
    if not rutas:
        sys.exit(f"No hay ningún .xlsx en {a.matrices}")

    print(f"Protocolos · P2342 Arqueros — corte {corte}")
    print(f"{len(rutas)} archivo(s) en {a.matrices}\n")

    DISC, fechas = {}, {}
    for r in rutas:
        d, fecha, calc, cache = leer_matriz(r)
        if not d:
            aviso(f"«{Path(r).name}» no parece una matriz de protocolos: se ignora.")
            continue
        if d in DISC:
            aviso(f"Hay dos matrices de {d}; me quedo con «{Path(r).name}».")
        DISC[d] = calc
        fechas[d] = fecha
        print(f"  {d:5s} {Path(r).name[:58]:58s} matriz al {fecha or '¿?'}")
        if a.verificar_cache:
            for sig in sorted(calc):
                c, h = calc[sig], cache.get(sig, {})
                mio = " ".join(f"{k}={c[k]}" for k in ESTADOS)
                suyo = " ".join(f"{k}={h.get(k, 0)}" for k in ESTADOS)
                marca = "OK" if mio == suyo else "⚠ caché desfasada"
                print(f"        {sig:5s} recalc {mio}")
                if marca != "OK":
                    print(f"        {'':5s} libro  {suyo}   {marca}")

    faltan = {m for m, _ in MAPA.values()} - set(DISC)
    if faltan:
        sys.exit(f"\nFaltan matrices: {', '.join(sorted(faltan))}. "
                 f"El proyecto se carga completo o no se carga.")

    datos = {}
    for nid, (mtz, siglas) in MAPA.items():
        t = {k: 0 for k in ESTADOS}
        for s in siglas:
            src = DISC[mtz].get(s)
            if src is None:
                aviso(f"{nid}: la matriz {mtz} no trae la columna «{s}».")
                continue
            for k in ESTADOS:
                t[k] += src[k]
        datos[nid] = t
    for nid, v in CONGELADOS.items():
        datos[nid] = dict(v, N=0)
        aviso(f"{nid}: sin matriz desde el 29-06-2026 — conserva "
              f"{v['S'] + v['C'] + v['P'] + v['AP'] + v['AE']} protocolos y no suma "
              f"punto nuevo al historial.")

    print()
    tot = {k: 0 for k in ESTADOS}
    for nid in sorted(datos):
        v = datos[nid]
        u = v["S"] + v["C"] + v["P"] + v["AP"] + v["AE"]
        for k in ESTADOS:
            tot[k] += v[k]
        print(f"  {nid:18s} S={v['S']:5d} C={v['C']:6d} P={v['P']:3d} "
              f"AP={v['AP']:3d} AE={v['AE']:5d}  universo {u:6d}  KPI {kpi_pct(v):5.2f}%")
    U = tot["S"] + tot["C"] + tot["P"] + tot["AP"] + tot["AE"]
    print(f"\n  {'TOTAL P2342':18s} S={tot['S']:5d} C={tot['C']:6d} P={tot['P']:3d} "
          f"AP={tot['AP']:3d} AE={tot['AE']:5d}  universo {U:6d}  KPI {kpi_pct(tot):5.2f}%")

    if a.seco:
        _fin()
        return

    html = Path(a.panel).read_text(encoding="utf-8")
    antes = len(html)
    html = inyectar(html, datos, corte, fechas)
    html, k, disc = inyectar_kpi(html, datos, corte)
    Path(a.panel).write_text(html, encoding="utf-8")
    print(f"\n  panel escrito: {a.panel}  ({antes:,} → {len(html):,} bytes)")
    print(f"  KPI  P2342 {k['P2342']}%  ·  P2407 {k['P2407']}%  ·  P2416 {k['P2416']}%  ·  CORP {k['CORP']}%")
    print("  por disciplina: " + ", ".join(f"{a} {b}%" for a, b in disc.items()))
    _fin()


def _fin():
    if AVISOS:
        print("\n⚠ AVISOS")
        for t in AVISOS:
            print(f"  · {t}")
    else:
        print("\nSin avisos.")


if __name__ == "__main__":
    main()
